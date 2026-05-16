#!/usr/bin/env python3
"""Compare human mean scores against LLM-as-a-judge scores."""

from __future__ import annotations

import argparse
import json
import math
import sys
from collections import Counter
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from analyze_human_eval_results import ANNOTATOR_SHEETS
from analyze_human_eval_results import CRITERIA
from analyze_human_eval_results import mean
from analyze_human_eval_results import pearson
from analyze_human_eval_results import read_annotator_sheet
from analyze_human_eval_results import rounded_or_none
from analyze_human_eval_results import sample_sort_key
from analyze_human_eval_results import spearman
from analyze_human_eval_results import validate_annotators
from analyze_human_eval_results import weighted_kappa
from analyze_human_eval_results import write_csv


DEFAULT_INPUT_XLSX = "human_eval/QUARE_Human_Evaluation_Collection_with_LLM_Judge.xlsx"
DEFAULT_OUTPUT_DIR = "human_eval/current_analysis_synced"
LLM_SHEET_CANDIDATES = ("LLM_Judge_Scores", "LLM-Judge(Hidden to Annotator)")


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare human and LLM judge scores.")
    parser.add_argument("--input-xlsx", default=DEFAULT_INPUT_XLSX)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    workbook = load_workbook(args.input_xlsx, data_only=True)
    annot_001 = read_annotator_sheet(workbook, ANNOTATOR_SHEETS[0])
    annot_002 = read_annotator_sheet(workbook, ANNOTATOR_SHEETS[1])
    sample_ids = validate_annotators(annot_001, annot_002)
    llm_scores = read_llm_scores(workbook)

    missing = [sample_id for sample_id in sample_ids if sample_id not in llm_scores]
    if missing:
        raise ValueError(f"Missing LLM scores for Sample_IDs: {missing}")

    alignment = build_alignment_rows(annot_001, annot_002, llm_scores, sample_ids)
    framework_summary = build_framework_rows(annot_001, annot_002, llm_scores, sample_ids)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "human_vs_llm_alignment.csv", alignment)
    write_csv(output_dir / "human_vs_llm_framework_summary.csv", framework_summary)
    (output_dir / "human_vs_llm_alignment.json").write_text(
        json.dumps(
            {
                "metadata": {
                    "input_xlsx": args.input_xlsx,
                    "n_requirements": len(sample_ids),
                    "criteria": [criterion for criterion, _ in CRITERIA],
                },
                "alignment": alignment,
                "framework_summary": framework_summary,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    print(f"Compared {len(sample_ids)} requirements.")
    print(f"Wrote {output_dir / 'human_vs_llm_alignment.csv'}")
    print(f"Wrote {output_dir / 'human_vs_llm_framework_summary.csv'}")
    print(f"Wrote {output_dir / 'human_vs_llm_alignment.json'}")
    return 0


def read_llm_scores(workbook: Any) -> dict[str, dict[str, Any]]:
    sheet_name = next((name for name in LLM_SHEET_CANDIDATES if name in workbook.sheetnames), "")
    if not sheet_name:
        raise ValueError(f"Could not find any LLM score sheet: {LLM_SHEET_CANDIDATES}")
    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    required = ("Sample_ID", "Framework", *(criterion for criterion, _ in CRITERIA))
    missing = [header for header in required if header not in index]
    if missing:
        raise ValueError(f"{sheet_name} is missing columns: {missing}")

    scores: dict[str, dict[str, Any]] = {}
    for row_idx in range(2, sheet.max_row + 1):
        sample_id = str(sheet.cell(row_idx, index["Sample_ID"]).value or "").strip()
        if not sample_id:
            continue
        row: dict[str, Any] = {
            "Sample_ID": sample_id,
            "Framework": str(sheet.cell(row_idx, index["Framework"]).value or "").strip(),
        }
        for criterion, _ in CRITERIA:
            value = sheet.cell(row_idx, index[criterion]).value
            if value in {None, ""}:
                raise ValueError(f"Missing LLM score for {sample_id} / {criterion}")
            score = int(value)
            if score not in {1, 2, 3, 4, 5}:
                raise ValueError(f"Out-of-range LLM score for {sample_id} / {criterion}: {score}")
            row[criterion] = score
        scores[sample_id] = row
    return scores


def build_alignment_rows(
    annot_001: dict[str, dict[str, Any]],
    annot_002: dict[str, dict[str, Any]],
    llm_scores: dict[str, dict[str, Any]],
    sample_ids: list[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    all_human: list[float] = []
    all_llm: list[int] = []
    all_human_rounded: list[int] = []

    for criterion, _ in CRITERIA:
        human = [
            (annot_001[sample_id][criterion] + annot_002[sample_id][criterion]) / 2.0
            for sample_id in sample_ids
        ]
        llm = [llm_scores[sample_id][criterion] for sample_id in sample_ids]
        human_rounded = [round_half_up(value) for value in human]
        rows.append(alignment_row(criterion, human, llm, human_rounded))
        all_human.extend(human)
        all_llm.extend(llm)
        all_human_rounded.extend(human_rounded)

    rows.append(alignment_row("OVERALL_ALL_CELLS", all_human, all_llm, all_human_rounded))
    return rows


def alignment_row(
    label: str,
    human: list[float],
    llm: list[int],
    human_rounded: list[int],
) -> dict[str, Any]:
    diffs = [abs(left - right) for left, right in zip(human, llm, strict=True)]
    rounded_diffs = [
        abs(left - right) for left, right in zip(human_rounded, llm, strict=True)
    ]
    n = len(human)
    return {
        "criterion": label,
        "n": n,
        "mean_human": round(mean(human), 6),
        "mean_llm": round(mean(llm), 6),
        "mean_abs_error": round(mean(diffs), 6),
        "exact_agreement_pct_vs_rounded_human": round(
            100.0 * sum(diff == 0 for diff in rounded_diffs) / n,
            6,
        ),
        "within_1_pct_vs_rounded_human": round(
            100.0 * sum(diff <= 1 for diff in rounded_diffs) / n,
            6,
        ),
        "quadratic_weighted_kappa_vs_rounded_human": rounded_or_none(
            weighted_kappa(human_rounded, llm)
        ),
        "spearman": rounded_or_none(spearman(human, llm)),
        "pearson": rounded_or_none(pearson(human, llm)),
        "score_dist_human_rounded": json.dumps(dict(sorted(Counter(human_rounded).items()))),
        "score_dist_llm": json.dumps(dict(sorted(Counter(llm).items()))),
    }


def build_framework_rows(
    annot_001: dict[str, dict[str, Any]],
    annot_002: dict[str, dict[str, Any]],
    llm_scores: dict[str, dict[str, Any]],
    sample_ids: list[str],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[float]] = defaultdict(list)
    grouped_llm: dict[tuple[str, str], list[int]] = defaultdict(list)
    criterion_human: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    criterion_llm: dict[tuple[str, str, str], list[int]] = defaultdict(list)

    for sample_id in sorted(sample_ids, key=sample_sort_key):
        framework = llm_scores[sample_id]["Framework"]
        case_study = str(annot_001[sample_id]["Case_Study"])
        for criterion, _ in CRITERIA:
            human = (annot_001[sample_id][criterion] + annot_002[sample_id][criterion]) / 2.0
            llm = llm_scores[sample_id][criterion]
            grouped[(framework, "ALL")].append(human)
            grouped[(framework, case_study)].append(human)
            grouped_llm[(framework, "ALL")].append(llm)
            grouped_llm[(framework, case_study)].append(llm)
            criterion_human[(framework, "ALL", criterion)].append(human)
            criterion_human[(framework, case_study, criterion)].append(human)
            criterion_llm[(framework, "ALL", criterion)].append(llm)
            criterion_llm[(framework, case_study, criterion)].append(llm)

    rows: list[dict[str, Any]] = []
    for framework, case_study in sorted(
        grouped,
        key=lambda item: (item[0], item[1] != "ALL", item[1]),
    ):
        row: dict[str, Any] = {
            "framework": framework,
            "case_study": case_study,
            "n_scores": len(grouped[(framework, case_study)]),
            "n_requirements": int(len(grouped[(framework, case_study)]) / len(CRITERIA)),
            "mean_human_score": round(mean(grouped[(framework, case_study)]), 6),
            "mean_llm_score": round(mean(grouped_llm[(framework, case_study)]), 6),
            "mean_llm_minus_human": round(
                mean(grouped_llm[(framework, case_study)])
                - mean(grouped[(framework, case_study)]),
                6,
            ),
        }
        for criterion, _ in CRITERIA:
            human_values = criterion_human[(framework, case_study, criterion)]
            llm_values = criterion_llm[(framework, case_study, criterion)]
            row[f"human_{criterion}"] = round(mean(human_values), 6)
            row[f"llm_{criterion}"] = round(mean(llm_values), 6)
        rows.append(row)
    return rows


def round_half_up(value: float) -> int:
    return int(math.floor(value + 0.5))


if __name__ == "__main__":
    raise SystemExit(main())
