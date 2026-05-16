#!/usr/bin/env python3
"""Generate the QUARE human-evaluation data collection workbook."""

from __future__ import annotations

import argparse
import json
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.worksheet.datavalidation import DataValidation


DEFAULT_INPUT_XLSX = "human_eval/human_eval_annotation.xlsx"
DEFAULT_MAPPING_JSON = "human_eval/human_eval_mapping.json"
DEFAULT_OUTPUT_XLSX = "human_eval/QUARE_Human_Evaluation_Collection.xlsx"
ANNOTATOR_SHEETS = ("ANNOT-001", "ANNOT-002")
ANNOTATION_COLUMNS = (
    "Sample_ID",
    "Case_Study",
    "Requirement_Set",
    "Requirement_Text",
    "Unambiguous(1-5)",
    "Correctness(1-5)",
    "Verifiability(1-5)",
    "Set_Consistency(1-5)",
    "Set_Feasibility(1-5)",
    "Notes",
)
LLM_COLUMNS = (
    "Sample_ID",
    "Case_Study",
    "Requirement_Set",
    "Framework",
    "Unambiguous",
    "Correctness",
    "Verifiability",
    "Set_Consistency",
    "Set_Feasibility",
)
HUMAN_SCORE_COLUMNS = (
    "Unambiguous(1-5)",
    "Correctness(1-5)",
    "Verifiability(1-5)",
    "Set_Consistency(1-5)",
    "Set_Feasibility(1-5)",
)
LLM_SCORE_COLUMNS = (
    "Unambiguous",
    "Correctness",
    "Verifiability",
    "Set_Consistency",
    "Set_Feasibility",
)
LLM_FILL_COMMENT = (
    "TO BE FILLED: run LLM judge on these 82 requirements using the same "
    "evaluation prompt from the experiment."
)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Create the QUARE human evaluation collection workbook."
    )
    parser.add_argument("--input-xlsx", default=DEFAULT_INPUT_XLSX)
    parser.add_argument("--mapping-json", default=DEFAULT_MAPPING_JSON)
    parser.add_argument("--output-xlsx", default=DEFAULT_OUTPUT_XLSX)
    args = parser.parse_args()

    input_xlsx = Path(args.input_xlsx)
    mapping_json = Path(args.mapping_json)
    output_xlsx = Path(args.output_xlsx)

    sample_mapping = load_sample_mapping(mapping_json)
    workbook = load_workbook(input_xlsx)

    require_sheets(workbook, {"Rubric", "Case Context", "Annotation"})
    annotation_rows = read_annotation_rows(workbook["Annotation"])
    validate_mapping_coverage(annotation_rows, sample_mapping)

    workbook["Rubric"].title = "Instructions"
    workbook["Case Context"].title = "Case_Context"
    annotation_sheet = workbook["Annotation"]
    annotation_sheet.title = ANNOTATOR_SHEETS[0]

    annotator_two = workbook.copy_worksheet(annotation_sheet)
    annotator_two.title = ANNOTATOR_SHEETS[1]

    for sheet_name in ANNOTATOR_SHEETS:
        prepare_annotator_sheet(workbook[sheet_name])

    llm_sheet = workbook.create_sheet("LLM_Judge_Scores")
    populate_llm_sheet(llm_sheet, annotation_rows, sample_mapping)
    style_llm_sheet(llm_sheet)

    workbook._sheets = [
        workbook["Instructions"],
        workbook["Case_Context"],
        workbook[ANNOTATOR_SHEETS[0]],
        workbook[ANNOTATOR_SHEETS[1]],
        workbook["LLM_Judge_Scores"],
    ]
    workbook.active = 0

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)

    print(f"Wrote {output_xlsx}")
    print(f"Requirements per annotator: {len(annotation_rows)}")
    return 0


def load_sample_mapping(mapping_json: Path) -> dict[str, dict[str, Any]]:
    payload = json.loads(mapping_json.read_text(encoding="utf-8"))
    samples = payload.get("samples", [])
    if not isinstance(samples, list):
        raise ValueError("Mapping JSON must contain a list at key 'samples'.")

    by_id: dict[str, dict[str, Any]] = {}
    for sample in samples:
        if not isinstance(sample, dict):
            continue
        sample_id = str(sample.get("sample_id", "")).strip()
        if not sample_id:
            continue
        if sample_id in by_id:
            raise ValueError(f"Duplicate sample_id in mapping JSON: {sample_id}")
        by_id[sample_id] = sample
    return by_id


def require_sheets(workbook: Any, required: set[str]) -> None:
    missing = sorted(required - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"Input workbook is missing required sheets: {missing}")


def read_annotation_rows(sheet: Any) -> list[dict[str, Any]]:
    headers = [sheet.cell(1, col_idx).value for col_idx in range(1, sheet.max_column + 1)]
    missing = [column for column in ANNOTATION_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Annotation sheet is missing columns: {missing}")

    col_index = {header: headers.index(header) + 1 for header in ANNOTATION_COLUMNS}
    rows: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for row_idx in range(2, sheet.max_row + 1):
        sample_id = sheet.cell(row_idx, col_index["Sample_ID"]).value
        if sample_id is None or not str(sample_id).strip():
            continue
        row = {
            column: sheet.cell(row_idx, col_index[column]).value
            for column in ANNOTATION_COLUMNS
        }
        row["Sample_ID"] = str(row["Sample_ID"]).strip()
        if row["Sample_ID"] in seen_ids:
            raise ValueError(f"Duplicate Sample_ID in annotation sheet: {row['Sample_ID']}")
        seen_ids.add(row["Sample_ID"])
        rows.append(row)
    return rows


def validate_mapping_coverage(
    annotation_rows: list[dict[str, Any]],
    sample_mapping: dict[str, dict[str, Any]],
) -> None:
    missing = [
        row["Sample_ID"]
        for row in annotation_rows
        if row["Sample_ID"] not in sample_mapping
    ]
    if missing:
        raise ValueError(f"Mapping JSON has no entries for Sample_IDs: {missing}")


def prepare_annotator_sheet(sheet: Any) -> None:
    clear_human_response_columns(sheet)
    sheet.freeze_panes = "A2"
    add_score_validation(sheet, HUMAN_SCORE_COLUMNS)
    style_annotation_like_sheet(sheet, ANNOTATION_COLUMNS)


def clear_human_response_columns(sheet: Any) -> None:
    headers = [sheet.cell(1, col_idx).value for col_idx in range(1, sheet.max_column + 1)]
    col_index = {header: headers.index(header) + 1 for header in ANNOTATION_COLUMNS}
    columns_to_clear = [*HUMAN_SCORE_COLUMNS, "Notes"]
    for row_idx in range(2, sheet.max_row + 1):
        for column_name in columns_to_clear:
            cell = sheet.cell(row_idx, col_index[column_name])
            cell.value = None
            cell.comment = None


def populate_llm_sheet(
    sheet: Any,
    annotation_rows: list[dict[str, Any]],
    sample_mapping: dict[str, dict[str, Any]],
) -> None:
    for col_idx, header in enumerate(LLM_COLUMNS, start=1):
        cell = sheet.cell(1, col_idx, header)
        if header in LLM_SCORE_COLUMNS:
            cell.comment = Comment(LLM_FILL_COMMENT, "Codex")

    for row_idx, row in enumerate(annotation_rows, start=2):
        sample = sample_mapping[row["Sample_ID"]]
        values = {
            "Sample_ID": row["Sample_ID"],
            "Case_Study": row["Case_Study"],
            "Requirement_Set": row["Requirement_Set"],
            "Framework": sample.get("framework", ""),
            "Unambiguous": None,
            "Correctness": None,
            "Verifiability": None,
            "Set_Consistency": None,
            "Set_Feasibility": None,
        }
        for col_idx, header in enumerate(LLM_COLUMNS, start=1):
            sheet.cell(row_idx, col_idx, values[header])


def style_llm_sheet(sheet: Any) -> None:
    sheet.freeze_panes = "A2"
    style_annotation_like_sheet(sheet, LLM_COLUMNS, requirement_text_column=None)
    add_score_validation(sheet, LLM_SCORE_COLUMNS)


def add_score_validation(sheet: Any, score_columns: tuple[str, ...]) -> None:
    sheet.data_validations.dataValidation = []
    headers = [sheet.cell(1, col_idx).value for col_idx in range(1, sheet.max_column + 1)]
    header_to_col = {header: headers.index(header) + 1 for header in headers}

    validation = DataValidation(
        type="list",
        formula1='"1,2,3,4,5"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid score",
        error="Please choose an integer score from 1 to 5.",
    )
    sheet.add_data_validation(validation)
    for column_name in score_columns:
        col_idx = header_to_col[column_name]
        validation.add(f"{sheet.cell(2, col_idx).coordinate}:{sheet.cell(sheet.max_row, col_idx).coordinate}")


def style_annotation_like_sheet(
    sheet: Any,
    columns: tuple[str, ...],
    *,
    requirement_text_column: str | None = "Requirement_Text",
) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin_gray)

    for col_idx in range(1, len(columns) + 1):
        cell = sheet.cell(1, col_idx)
        cell.fill = copy(header_fill)
        cell.font = copy(header_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(border)

    header_to_col = {header: index + 1 for index, header in enumerate(columns)}
    score_names = set(HUMAN_SCORE_COLUMNS) | set(LLM_SCORE_COLUMNS)
    score_cols = [header_to_col[name] for name in columns if name in score_names]
    requirement_col = header_to_col.get(requirement_text_column) if requirement_text_column else None

    for row_idx in range(2, sheet.max_row + 1):
        for col_idx in range(1, len(columns) + 1):
            cell = sheet.cell(row_idx, col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=col_idx == requirement_col)
        for col_idx in score_cols:
            sheet.cell(row_idx, col_idx).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    set_column_widths(sheet, columns, requirement_col=requirement_col)


def set_column_widths(
    sheet: Any,
    columns: tuple[str, ...],
    *,
    requirement_col: int | None,
) -> None:
    fixed_widths = {
        "Sample_ID": 12,
        "Case_Study": 16,
        "Requirement_Set": 16,
        "Requirement_Text": 60,
        "Framework": 14,
        "Notes": 36,
    }
    score_names = set(HUMAN_SCORE_COLUMNS) | set(LLM_SCORE_COLUMNS)

    for col_idx, header in enumerate(columns, start=1):
        letter = sheet.cell(1, col_idx).column_letter
        if header in fixed_widths:
            width = fixed_widths[header]
        elif header in score_names:
            width = 18
        else:
            values = [
                str(sheet.cell(row_idx, col_idx).value or "")
                for row_idx in range(1, sheet.max_row + 1)
            ]
            width = min(max(max(len(value) for value in values) + 2, 10), 28)
        sheet.column_dimensions[letter].width = width

    if requirement_col is not None:
        for row_idx in range(2, sheet.max_row + 1):
            text = str(sheet.cell(row_idx, requirement_col).value or "")
            sheet.row_dimensions[row_idx].height = min(180, max(36, (len(text) // 70 + 1) * 18))


if __name__ == "__main__":
    raise SystemExit(main())
