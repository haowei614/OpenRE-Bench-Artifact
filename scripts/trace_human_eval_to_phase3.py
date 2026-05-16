#!/usr/bin/env python3
"""Trace human-evaluation workbook rows back to phase3 integrated model files."""

from __future__ import annotations

import argparse
import csv
import difflib
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from build_human_eval_workbook import format_raw_requirement_text
from build_human_eval_workbook import format_requirement_text
from build_human_eval_workbook import is_requirement_like
from build_human_eval_workbook import normalize_text


DEFAULT_INPUT_XLSX = "human_eval/QUARE_Human_Evaluation_Collection.xlsx"
DEFAULT_RUNS_DIR = "experiment_outputs/mare-iredev-quare/runs"
DEFAULT_OUTPUT_CSV = "human_eval/current_analysis_synced/human_eval_phase3_traceability.csv"
DEFAULT_OUTPUT_JSON = "human_eval/current_analysis_synced/human_eval_phase3_traceability.json"
ANNOTATION_SHEET = "ANNOT-001"
LLM_SHEET = "LLM_Judge_Scores"
PHASE3_FILENAME = "phase3_integrated_kaos_model.json"
RUN_RECORD_FILENAME = "run_record.json"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Check whether human-eval samples can be traced to phase3 JSON elements."
    )
    parser.add_argument("--input-xlsx", default=DEFAULT_INPUT_XLSX)
    parser.add_argument("--runs-dir", default=DEFAULT_RUNS_DIR)
    parser.add_argument("--output-csv", default=DEFAULT_OUTPUT_CSV)
    parser.add_argument("--output-json", default=DEFAULT_OUTPUT_JSON)
    args = parser.parse_args()

    workbook = load_workbook(args.input_xlsx, data_only=True)
    samples = read_samples(workbook)
    workbook_frameworks = read_workbook_frameworks(workbook)
    phase3_index, phase3_records = build_phase3_index(Path(args.runs_dir))
    trace_rows = build_trace_rows(samples, workbook_frameworks, phase3_index, phase3_records)

    output_csv = Path(args.output_csv)
    output_json = Path(args.output_json)
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    write_csv(output_csv, trace_rows)
    output_json.write_text(
        json.dumps(
            {
                "metadata": {
                    "input_xlsx": args.input_xlsx,
                    "runs_dir": args.runs_dir,
                    "n_samples": len(samples),
                    "n_phase3_records": len(phase3_records),
                    "n_exact_matches": sum(row["trace_status"] == "MATCHED" for row in trace_rows),
                    "n_unmatched": sum(row["trace_status"] != "MATCHED" for row in trace_rows),
                    "n_framework_consistent": sum(
                        row["framework_consistent"] == "yes" for row in trace_rows
                    ),
                    "n_framework_mismatch": sum(
                        row["framework_consistent"] == "no" for row in trace_rows
                    ),
                },
                "trace_rows": trace_rows,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    print(f"Samples in workbook: {len(samples)}")
    print(f"Phase3 element texts indexed: {len(phase3_records)}")
    print(f"Exact cleaned-text matches: {sum(row['trace_status'] == 'MATCHED' for row in trace_rows)}")
    print(
        "Framework-consistent matches: "
        f"{sum(row['framework_consistent'] == 'yes' for row in trace_rows)}"
    )
    print(
        "Framework mismatches: "
        f"{sum(row['framework_consistent'] == 'no' for row in trace_rows)}"
    )
    print(f"Wrote {output_csv}")
    print(f"Wrote {output_json}")
    return 0


def read_samples(workbook: Any) -> list[dict[str, str]]:
    sheet = workbook[ANNOTATION_SHEET]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    required = ("Sample_ID", "Case_Study", "Requirement_Set", "Requirement_Text")
    missing = [header for header in required if header not in index]
    if missing:
        raise ValueError(f"{ANNOTATION_SHEET} missing columns: {missing}")

    samples: list[dict[str, str]] = []
    seen: set[str] = set()
    for row_idx in range(2, sheet.max_row + 1):
        sample_id = str(sheet.cell(row_idx, index["Sample_ID"]).value or "").strip()
        if not sample_id:
            continue
        if sample_id in seen:
            raise ValueError(f"Duplicate Sample_ID in {ANNOTATION_SHEET}: {sample_id}")
        seen.add(sample_id)
        samples.append(
            {
                "Sample_ID": sample_id,
                "Case_Study": str(sheet.cell(row_idx, index["Case_Study"]).value or "").strip(),
                "Requirement_Set": str(
                    sheet.cell(row_idx, index["Requirement_Set"]).value or ""
                ).strip(),
                "Requirement_Text": str(
                    sheet.cell(row_idx, index["Requirement_Text"]).value or ""
                ).strip(),
            }
        )
    return samples


def read_workbook_frameworks(workbook: Any) -> dict[str, str]:
    if LLM_SHEET not in workbook.sheetnames:
        return {}
    sheet = workbook[LLM_SHEET]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    if "Sample_ID" not in index or "Framework" not in index:
        return {}

    frameworks: dict[str, str] = {}
    for row_idx in range(2, sheet.max_row + 1):
        sample_id = str(sheet.cell(row_idx, index["Sample_ID"]).value or "").strip()
        if not sample_id:
            continue
        frameworks[sample_id] = str(sheet.cell(row_idx, index["Framework"]).value or "").strip()
    return frameworks


def build_phase3_index(
    runs_dir: Path,
) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    records: list[dict[str, Any]] = []
    index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for run_record_path in sorted(runs_dir.glob(f"*/{RUN_RECORD_FILENAME}")):
        run_record = load_json(run_record_path)
        if not isinstance(run_record, dict):
            continue
        phase3_path = resolve_phase3_path(run_record, run_record_path.parent)
        if phase3_path is None or not phase3_path.exists():
            continue
        phase3 = load_json(phase3_path)
        elements = phase3.get("gsn_elements", []) if isinstance(phase3, dict) else []
        if not isinstance(elements, list):
            continue
        for element in elements:
            if not isinstance(element, dict):
                continue
            raw_text = format_raw_requirement_text(element)
            cleaned_text = format_requirement_text(element)
            if not cleaned_text:
                continue
            properties = element.get("properties", {})
            if not isinstance(properties, dict):
                properties = {}
            record = {
                "source_file": str(phase3_path),
                "run_id": str(run_record.get("run_id", "")).strip(),
                "source_framework": str(run_record.get("system", "")).strip().lower(),
                "case_study": str(run_record.get("case_id", "")).strip(),
                "setting": str(run_record.get("setting", "")).strip(),
                "seed": str(run_record.get("seed", "")).strip(),
                "source_requirement_id": str(element.get("id", "")).strip(),
                "source_requirement_name": str(element.get("name", "")).strip(),
                "source_element_type": (
                    str(properties.get("original_kaos_type", "")).strip()
                    or str(element.get("gsn_type", "")).strip()
                ),
                "quality_attribute": str(element.get("quality_attribute", "")).strip(),
                "stakeholder": str(element.get("stakeholder", "")).strip(),
                "is_requirement_like": is_requirement_like(
                    element,
                    requirement_text=cleaned_text,
                    include_goals=True,
                ),
                "raw_requirement_text": raw_text,
                "cleaned_requirement_text": cleaned_text,
            }
            records.append(record)
            index[normalize_text(cleaned_text)].append(record)
    return index, records


def build_trace_rows(
    samples: list[dict[str, str]],
    workbook_frameworks: dict[str, str],
    phase3_index: dict[str, list[dict[str, Any]]],
    phase3_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    trace_rows: list[dict[str, Any]] = []
    for sample in samples:
        sample_id = sample["Sample_ID"]
        workbook_framework = workbook_frameworks.get(sample_id, "")
        candidates = phase3_index.get(normalize_text(sample["Requirement_Text"]), [])
        same_case = [item for item in candidates if item["case_study"] == sample["Case_Study"]]
        same_framework = [
            item for item in same_case if item["source_framework"] == workbook_framework
        ]
        preferred = choose_preferred_match(same_framework or same_case or candidates)
        best_fuzzy = choose_best_fuzzy_match(sample, workbook_framework, phase3_records)
        matched = preferred is not None
        source_frameworks = sorted({item["source_framework"] for item in same_case or candidates})
        framework_consistent = ""
        if matched and workbook_framework:
            framework_consistent = "yes" if same_framework else "no"

        trace_rows.append(
            {
                "Sample_ID": sample_id,
                "Case_Study": sample["Case_Study"],
                "Requirement_Set": sample["Requirement_Set"],
                "Workbook_Framework": workbook_framework,
                "Requirement_Text": sample["Requirement_Text"],
                "trace_status": "MATCHED" if matched else "UNMATCHED",
                "n_exact_text_matches": len(candidates),
                "n_same_case_matches": len(same_case),
                "n_same_framework_matches": len(same_framework),
                "source_frameworks_for_same_case": ";".join(source_frameworks),
                "framework_consistent": framework_consistent,
                "matched_source_framework": preferred.get("source_framework", "") if preferred else "",
                "matched_case_study": preferred.get("case_study", "") if preferred else "",
                "matched_setting": preferred.get("setting", "") if preferred else "",
                "matched_seed": preferred.get("seed", "") if preferred else "",
                "matched_run_id": preferred.get("run_id", "") if preferred else "",
                "matched_source_requirement_id": preferred.get("source_requirement_id", "")
                if preferred
                else "",
                "matched_source_requirement_name": preferred.get("source_requirement_name", "")
                if preferred
                else "",
                "matched_source_element_type": preferred.get("source_element_type", "")
                if preferred
                else "",
                "matched_quality_attribute": preferred.get("quality_attribute", "")
                if preferred
                else "",
                "matched_stakeholder": preferred.get("stakeholder", "") if preferred else "",
                "matched_is_requirement_like": preferred.get("is_requirement_like", "")
                if preferred
                else "",
                "matched_source_file": preferred.get("source_file", "") if preferred else "",
                "matched_raw_requirement_text": preferred.get("raw_requirement_text", "")
                if preferred
                else "",
                "best_fuzzy_ratio_same_case": best_fuzzy.get("ratio", ""),
                "best_fuzzy_source_framework": best_fuzzy.get("source_framework", ""),
                "best_fuzzy_setting": best_fuzzy.get("setting", ""),
                "best_fuzzy_seed": best_fuzzy.get("seed", ""),
                "best_fuzzy_run_id": best_fuzzy.get("run_id", ""),
                "best_fuzzy_source_requirement_id": best_fuzzy.get("source_requirement_id", ""),
                "best_fuzzy_source_file": best_fuzzy.get("source_file", ""),
                "best_fuzzy_cleaned_requirement_text": best_fuzzy.get(
                    "cleaned_requirement_text",
                    "",
                ),
            }
        )
    return trace_rows


def choose_preferred_match(candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not candidates:
        return None
    return sorted(
        candidates,
        key=lambda item: (
            item["setting"] != "negotiation_integration_verification",
            not bool(item["is_requirement_like"]),
            item["run_id"],
            item["source_requirement_id"],
        ),
    )[0]


def choose_best_fuzzy_match(
    sample: dict[str, str],
    workbook_framework: str,
    phase3_records: list[dict[str, Any]],
) -> dict[str, Any]:
    sample_text = normalize_text(sample["Requirement_Text"])
    same_case = [
        item for item in phase3_records if item["case_study"] == sample["Case_Study"]
    ]
    same_framework = [
        item for item in same_case if item["source_framework"] == workbook_framework
    ]
    candidates = same_framework or same_case
    if not candidates:
        return {}

    best: dict[str, Any] = {}
    best_score = -1.0
    for item in candidates:
        candidate_text = normalize_text(str(item["cleaned_requirement_text"]))
        score = difflib.SequenceMatcher(None, sample_text, candidate_text).ratio()
        if score > best_score:
            best_score = score
            best = dict(item)
            best["ratio"] = round(best_score, 6)
    return best


def resolve_phase3_path(run_record: dict[str, Any], run_dir: Path) -> Path | None:
    artifact_paths = run_record.get("artifact_paths", {})
    if isinstance(artifact_paths, dict):
        raw_path = str(artifact_paths.get(PHASE3_FILENAME, "")).strip()
        if raw_path:
            candidate = Path(raw_path)
            if candidate.exists():
                return candidate
            relative_candidate = run_dir / candidate.name
            if relative_candidate.exists():
                return relative_candidate
    candidate = run_dir / PHASE3_FILENAME
    if candidate.exists():
        return candidate
    return None


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


if __name__ == "__main__":
    raise SystemExit(main())
