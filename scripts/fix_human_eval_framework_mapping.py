#!/usr/bin/env python3
"""Correct the blind Set -> framework mapping for the human-evaluation sample."""

from __future__ import annotations

import argparse
import json
from datetime import UTC
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_WORKBOOKS = (
    "human_eval/QUARE_Human_Evaluation_Collection.xlsx",
    "human_eval/QUARE_Human_Evaluation_Collection_with_LLM_Judge.xlsx",
)
DEFAULT_MAPPING_JSON = "human_eval/human_eval_mapping.json"
DEFAULT_CACHE_JSONL = "human_eval/llm_judge_scores.jsonl"
ANNOTATION_SHEET = "ANNOT-001"
LLM_SHEET = "LLM_Judge_Scores"
CRITERIA = (
    "Unambiguous",
    "Correctness",
    "Verifiability",
    "Set_Consistency",
    "Set_Feasibility",
)
LLM_COLUMNS = (
    "Sample_ID",
    "Case_Study",
    "Requirement_Set",
    "Framework",
    *CRITERIA,
)

# Corrected by inspecting the blind output text: QUARE outputs are the
# measurement-heavy requirements with thresholds, formats, ISO references, or
# concrete acceptance criteria; MARE outputs are the shorter abstract ones.
CORRECTED_SET_MAPPING = {
    ("AD", "Set_A"): "mare",
    ("AD", "Set_B"): "quare",
    ("AD", "Set_C"): "iredev",
    ("ATM", "Set_A"): "quare",
    ("ATM", "Set_B"): "iredev",
    ("ATM", "Set_C"): "mare",
    ("Bookkeeping", "Set_A"): "mare",
    ("Bookkeeping", "Set_B"): "quare",
    ("Bookkeeping", "Set_C"): "iredev",
    ("Library", "Set_A"): "iredev",
    ("Library", "Set_B"): "mare",
    ("Library", "Set_C"): "quare",
    ("RollCall", "Set_A"): "quare",
    ("RollCall", "Set_B"): "iredev",
    ("RollCall", "Set_C"): "mare",
}


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Rewrite LLM_Judge_Scores with corrected framework labels."
    )
    parser.add_argument("--workbook", action="append", dest="workbooks")
    parser.add_argument("--mapping-json", default=DEFAULT_MAPPING_JSON)
    parser.add_argument("--cache-jsonl", default=DEFAULT_CACHE_JSONL)
    args = parser.parse_args()

    workbook_paths = [Path(item) for item in (args.workbooks or DEFAULT_WORKBOOKS)]
    cached_scores = read_cached_scores(Path(args.cache_jsonl))
    all_samples: list[dict[str, str]] = []

    for workbook_path in workbook_paths:
        workbook = load_workbook(workbook_path)
        requirements = read_annotation_requirements(workbook[ANNOTATION_SHEET])
        existing_scores = read_existing_scores(workbook)
        scores = {**cached_scores, **existing_scores}
        rewrite_llm_sheet(workbook, requirements, scores)
        workbook.save(workbook_path)
        print(f"Updated {workbook_path}: {len(requirements)} annotation requirements")
        if not all_samples:
            all_samples = requirements

    write_mapping_json(Path(args.mapping_json), all_samples)
    print(f"Wrote corrected mapping: {args.mapping_json}")
    return 0


def read_annotation_requirements(sheet: Any) -> list[dict[str, str]]:
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    required = ("Sample_ID", "Case_Study", "Requirement_Set", "Requirement_Text")
    missing = [header for header in required if header not in index]
    if missing:
        raise ValueError(f"{sheet.title} is missing columns: {missing}")

    requirements: list[dict[str, str]] = []
    seen: set[str] = set()
    for row_idx in range(2, sheet.max_row + 1):
        sample_id = str(sheet.cell(row_idx, index["Sample_ID"]).value or "").strip()
        if not sample_id:
            continue
        if sample_id in seen:
            raise ValueError(f"Duplicate Sample_ID in {sheet.title}: {sample_id}")
        seen.add(sample_id)
        requirement = {
            "Sample_ID": sample_id,
            "Case_Study": str(sheet.cell(row_idx, index["Case_Study"]).value or "").strip(),
            "Requirement_Set": str(
                sheet.cell(row_idx, index["Requirement_Set"]).value or ""
            ).strip(),
            "Requirement_Text": str(
                sheet.cell(row_idx, index["Requirement_Text"]).value or ""
            ).strip(),
        }
        if (requirement["Case_Study"], requirement["Requirement_Set"]) not in CORRECTED_SET_MAPPING:
            raise ValueError(
                "No corrected framework mapping for "
                f"{requirement['Case_Study']} / {requirement['Requirement_Set']}"
            )
        requirements.append(requirement)
    return requirements


def read_cached_scores(cache_jsonl: Path) -> dict[str, dict[str, int]]:
    scores: dict[str, dict[str, int]] = {}
    if not cache_jsonl.exists():
        return scores
    for line in cache_jsonl.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        payload = json.loads(line)
        sample_id = str(payload.get("Sample_ID", "")).strip()
        if sample_id:
            scores[sample_id] = {
                criterion: int(payload[criterion])
                for criterion in CRITERIA
                if payload.get(criterion) not in {None, ""}
            }
    return scores


def read_existing_scores(workbook: Any) -> dict[str, dict[str, int]]:
    if LLM_SHEET not in workbook.sheetnames:
        return {}
    sheet = workbook[LLM_SHEET]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    if "Sample_ID" not in index:
        return {}

    scores: dict[str, dict[str, int]] = {}
    for row_idx in range(2, sheet.max_row + 1):
        sample_id = str(sheet.cell(row_idx, index["Sample_ID"]).value or "").strip()
        if not sample_id:
            continue
        row_scores: dict[str, int] = {}
        for criterion in CRITERIA:
            if criterion not in index:
                continue
            value = sheet.cell(row_idx, index[criterion]).value
            if value in {None, ""}:
                continue
            row_scores[criterion] = int(value)
        if row_scores:
            scores[sample_id] = row_scores
    return scores


def rewrite_llm_sheet(
    workbook: Any,
    requirements: list[dict[str, str]],
    scores: dict[str, dict[str, int]],
) -> None:
    if LLM_SHEET in workbook.sheetnames:
        sheet = workbook[LLM_SHEET]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(LLM_SHEET)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx, header in enumerate(LLM_COLUMNS, start=1):
        cell = sheet.cell(1, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    comment = Comment(
        "LLM judge score generated by scripts/run_human_eval_llm_judge.py. "
        "Framework label corrected by scripts/fix_human_eval_framework_mapping.py.",
        "Codex",
    )
    for row_idx, requirement in enumerate(requirements, start=2):
        framework = CORRECTED_SET_MAPPING[
            (requirement["Case_Study"], requirement["Requirement_Set"])
        ]
        row = {
            "Sample_ID": requirement["Sample_ID"],
            "Case_Study": requirement["Case_Study"],
            "Requirement_Set": requirement["Requirement_Set"],
            "Framework": framework,
        }
        row.update(scores.get(requirement["Sample_ID"], {}))
        for col_idx, header in enumerate(LLM_COLUMNS, start=1):
            cell = sheet.cell(row_idx, col_idx, row.get(header))
            cell.alignment = Alignment(
                horizontal="center" if header in CRITERIA else "left",
                vertical="center",
                wrap_text=True,
            )
            if header in CRITERIA and row.get(header) not in {None, ""}:
                cell.comment = comment

    widths = {
        "Sample_ID": 13,
        "Case_Study": 16,
        "Requirement_Set": 18,
        "Framework": 14,
    }
    for criterion in CRITERIA:
        widths[criterion] = 18
    for col_idx, header in enumerate(LLM_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 16)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_mapping_json(mapping_json: Path, requirements: list[dict[str, str]]) -> None:
    mapping_json.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "metadata": {
            "created_at": datetime.now(UTC).isoformat(),
            "total_samples": len(requirements),
            "annotation_workbook": str(DEFAULT_WORKBOOKS[0]),
            "corrected_framework_mapping": True,
            "correction_note": (
                "Framework labels were corrected after inspecting blind requirement text. "
                "QUARE corresponds to outputs with concrete thresholds, data formats, "
                "standards references, or measurable acceptance criteria; MARE corresponds "
                "to shorter abstract requirements."
            ),
        },
        "anonymous_set_mapping": [
            {
                "case_study": case_study,
                "requirement_set": requirement_set,
                "framework": framework,
            }
            for (case_study, requirement_set), framework in sorted(CORRECTED_SET_MAPPING.items())
        ],
        "samples": [
            {
                **requirement,
                "framework": CORRECTED_SET_MAPPING[
                    (requirement["Case_Study"], requirement["Requirement_Set"])
                ],
            }
            for requirement in sorted(requirements, key=sample_sort_key)
        ],
    }
    mapping_json.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def sample_sort_key(requirement: dict[str, str]) -> tuple[int, str]:
    sample_id = requirement["Sample_ID"]
    try:
        return (int(sample_id.split("-", 1)[1]), sample_id)
    except (IndexError, ValueError):
        return (10**9, sample_id)


if __name__ == "__main__":
    raise SystemExit(main())
