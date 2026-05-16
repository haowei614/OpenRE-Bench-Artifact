#!/usr/bin/env python3
"""Match current human-evaluation collection rows to exported Phase 3 requirements."""

from __future__ import annotations

import argparse
import csv
import difflib
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from build_human_eval_workbook import normalize_text


DEFAULT_COLLECTION_XLSX = "human_eval/QUARE_Human_Evaluation_Collection_with_LLM_Judge.xlsx"
DEFAULT_PHASE3_JSON = "human_eval/phase3_requirements_all.json"
DEFAULT_OUTPUT_CSV = "human_eval/current_analysis_synced/collection_to_phase3_matches.csv"
DEFAULT_OUTPUT_JSON = "human_eval/current_analysis_synced/collection_to_phase3_matches.json"
ANNOTATION_SHEET = "ANNOT-001"
LLM_SHEET = "LLM_Judge_Scores"
CRITERIA = (
    "Unambiguous",
    "Correctness",
    "Verifiability",
    "Set_Consistency",
    "Set_Feasibility",
)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Compare 66 collection rows against exported Phase 3 requirement rows."
    )
    parser.add_argument("--collection-xlsx", default=DEFAULT_COLLECTION_XLSX)
    parser.add_argument("--phase3-json", default=DEFAULT_PHASE3_JSON)
    parser.add_argument("--output-csv", default=DEFAULT_OUTPUT_CSV)
    parser.add_argument("--output-json", default=DEFAULT_OUTPUT_JSON)
    args = parser.parse_args()

    collection = read_collection(Path(args.collection_xlsx))
    phase3 = json.loads(Path(args.phase3_json).read_text(encoding="utf-8"))
    if not isinstance(phase3, list):
        raise ValueError(f"{args.phase3_json} must contain a JSON list")

    rows = match_rows(collection, phase3)
    output_csv = Path(args.output_csv)
    output_json = Path(args.output_json)
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    write_csv(output_csv, rows)
    output_json.write_text(
        json.dumps(
            {
                "metadata": build_metadata(rows),
                "matches": rows,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    metadata = build_metadata(rows)
    print(f"Collection rows: {metadata['n_collection_rows']}")
    print(f"Exact normalized matches: {metadata['n_exact_matches']}")
    print(f"Exact same-framework matches: {metadata['n_exact_same_framework_matches']}")
    print(f"Fuzzy >= 0.90: {metadata['n_fuzzy_ge_0_90']}")
    print(f"Fuzzy >= 0.80: {metadata['n_fuzzy_ge_0_80']}")
    print(f"Fuzzy >= 0.70: {metadata['n_fuzzy_ge_0_70']}")
    print(f"Wrote {output_csv}")
    print(f"Wrote {output_json}")
    return 0


def read_collection(collection_xlsx: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(collection_xlsx, data_only=True)
    frameworks = read_frameworks(workbook)
    sheet = workbook[ANNOTATION_SHEET]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    required = ("Sample_ID", "Case_Study", "Requirement_Set", "Requirement_Text")
    missing = [header for header in required if header not in index]
    if missing:
        raise ValueError(f"{ANNOTATION_SHEET} missing columns: {missing}")

    rows: list[dict[str, Any]] = []
    for row_idx in range(2, sheet.max_row + 1):
        sample_id = str(sheet.cell(row_idx, index["Sample_ID"]).value or "").strip()
        if not sample_id:
            continue
        rows.append(
            {
                "Collection_Sample_ID": sample_id,
                "Case_Study": str(sheet.cell(row_idx, index["Case_Study"]).value or "").strip(),
                "Requirement_Set": str(
                    sheet.cell(row_idx, index["Requirement_Set"]).value or ""
                ).strip(),
                "Workbook_Framework": frameworks.get(sample_id, ""),
                "Requirement_Text": str(
                    sheet.cell(row_idx, index["Requirement_Text"]).value or ""
                ).strip(),
            }
        )
    return rows


def read_frameworks(workbook: Any) -> dict[str, str]:
    if LLM_SHEET not in workbook.sheetnames:
        return {}
    sheet = workbook[LLM_SHEET]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    if "Sample_ID" not in index or "Framework" not in index:
        return {}
    frameworks: dict[str, str] = {}
    for row_idx in range(2, sheet.max_row + 1):
        sample_id = str(sheet.cell(row_idx, index["Sample_ID"]).value or "").strip()
        if sample_id:
            frameworks[sample_id] = str(sheet.cell(row_idx, index["Framework"]).value or "").strip()
    return frameworks


def match_rows(collection: list[dict[str, Any]], phase3: list[dict[str, Any]]) -> list[dict[str, Any]]:
    exact_index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in phase3:
        exact_index[normalize_text(str(record.get("cleaned_requirement_text", "")))].append(record)

    output_rows: list[dict[str, Any]] = []
    for item in collection:
        normalized = normalize_text(str(item["Requirement_Text"]))
        exact_candidates = exact_index.get(normalized, [])
        exact_same_case = [
            record
            for record in exact_candidates
            if str(record.get("case_study", "")) == item["Case_Study"]
        ]
        exact_same_framework = [
            record
            for record in exact_same_case
            if str(record.get("framework", "")) == item["Workbook_Framework"]
        ]
        exact_preferred = choose_preferred(exact_same_framework or exact_same_case or exact_candidates)
        fuzzy_preferred = choose_best_fuzzy(item, phase3)
        row = {
            **item,
            "exact_match_status": "MATCHED" if exact_preferred else "UNMATCHED",
            "n_exact_text_matches": len(exact_candidates),
            "n_exact_same_case_matches": len(exact_same_case),
            "n_exact_same_framework_matches": len(exact_same_framework),
        }
        row.update(prefix_match("exact", exact_preferred))
        row.update(prefix_match("best_fuzzy", fuzzy_preferred))
        output_rows.append(row)
    return output_rows


def choose_preferred(candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not candidates:
        return None
    return sorted(
        candidates,
        key=lambda record: (
            str(record.get("setting", "")) != "negotiation_integration_verification",
            str(record.get("framework", "")),
            str(record.get("case_study", "")),
            int(record.get("seed", 0) or 0),
            str(record.get("source_requirement_id", "")),
        ),
    )[0]


def choose_best_fuzzy(item: dict[str, Any], phase3: list[dict[str, Any]]) -> dict[str, Any]:
    collection_text = normalize_text(str(item["Requirement_Text"]))
    same_case_framework = [
        record
        for record in phase3
        if str(record.get("case_study", "")) == item["Case_Study"]
        and str(record.get("framework", "")) == item["Workbook_Framework"]
    ]
    same_case = [
        record for record in phase3 if str(record.get("case_study", "")) == item["Case_Study"]
    ]
    candidates = same_case_framework or same_case or phase3
    best: dict[str, Any] = {}
    best_score = -1.0
    for record in candidates:
        phase3_text = normalize_text(str(record.get("cleaned_requirement_text", "")))
        score = difflib.SequenceMatcher(None, collection_text, phase3_text).ratio()
        if score > best_score:
            best_score = score
            best = dict(record)
            best["fuzzy_ratio"] = round(score, 6)
    return best


def prefix_match(prefix: str, record: dict[str, Any] | None) -> dict[str, Any]:
    fields = (
        "Sample_ID",
        "framework",
        "case_study",
        "setting",
        "seed",
        "run_id",
        "source_requirement_id",
        "source_file",
        "raw_requirement_text",
        "cleaned_requirement_text",
    )
    row = {f"{prefix}_{field}": "" for field in fields}
    if prefix == "best_fuzzy":
        row[f"{prefix}_ratio"] = ""
    if not record:
        return row
    for field in fields:
        row[f"{prefix}_{field}"] = record.get(field, "")
    if prefix == "best_fuzzy":
        row[f"{prefix}_ratio"] = record.get("fuzzy_ratio", "")
    return row


def build_metadata(rows: list[dict[str, Any]]) -> dict[str, Any]:
    ratios = [
        float(row["best_fuzzy_ratio"])
        for row in rows
        if str(row.get("best_fuzzy_ratio", "")).strip()
    ]
    return {
        "n_collection_rows": len(rows),
        "n_exact_matches": sum(row["exact_match_status"] == "MATCHED" for row in rows),
        "n_exact_same_framework_matches": sum(
            int(row["n_exact_same_framework_matches"]) > 0 for row in rows
        ),
        "n_fuzzy_ge_0_90": sum(ratio >= 0.90 for ratio in ratios),
        "n_fuzzy_ge_0_80": sum(ratio >= 0.80 for ratio in ratios),
        "n_fuzzy_ge_0_70": sum(ratio >= 0.70 for ratio in ratios),
        "min_fuzzy_ratio": round(min(ratios), 6) if ratios else None,
        "max_fuzzy_ratio": round(max(ratios), 6) if ratios else None,
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    raise SystemExit(main())
