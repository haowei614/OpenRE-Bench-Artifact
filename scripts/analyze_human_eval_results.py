#!/usr/bin/env python3
"""Analyze completed QUARE human-evaluation annotations."""

from __future__ import annotations

import argparse
import csv
import json
import math
from collections import Counter
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_INPUT_XLSX = "human_eval/QUARE_Human_Evaluation_Collection.xlsx"
DEFAULT_OUTPUT_DIR = "human_eval"
ANNOTATOR_SHEETS = ("ANNOT-001", "ANNOT-002")
LLM_SHEET_CANDIDATES = ("LLM_Judge_Scores", "LLM-Judge(Hidden to Annotator)")
CRITERIA = (
    ("Unambiguous", "Unambiguous(1-5)"),
    ("Correctness", "Correctness(1-5)"),
    ("Verifiability", "Verifiability(1-5)"),
    ("Set_Consistency", "Set_Consistency(1-5)"),
    ("Set_Feasibility", "Set_Feasibility(1-5)"),
)
SCORE_VALUES = (1, 2, 3, 4, 5)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Compute human-human agreement and descriptive summaries."
    )
    parser.add_argument("--input-xlsx", default=DEFAULT_INPUT_XLSX)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    workbook = load_workbook(args.input_xlsx, data_only=True)
    annot_001 = read_annotator_sheet(workbook, ANNOTATOR_SHEETS[0])
    annot_002 = read_annotator_sheet(workbook, ANNOTATOR_SHEETS[1])
    framework_by_id = read_framework_mapping(workbook)
    sample_ids = validate_annotators(annot_001, annot_002)

    agreement = build_agreement_summary(annot_001, annot_002, sample_ids)
    framework_summary = build_framework_summary(
        annot_001,
        annot_002,
        framework_by_id,
        sample_ids,
    )
    item_scores = build_item_scores(annot_001, annot_002, framework_by_id, sample_ids)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "human_eval_agreement_summary.csv", agreement)
    write_csv(output_dir / "human_eval_framework_summary.csv", framework_summary)
    write_csv(output_dir / "human_eval_item_level_scores.csv", item_scores)
    (output_dir / "human_eval_agreement_summary.json").write_text(
        json.dumps(
            {
                "metadata": {
                    "input_xlsx": args.input_xlsx,
                    "annotators": list(ANNOTATOR_SHEETS),
                    "n_requirements": len(sample_ids),
                    "criteria": [criterion for criterion, _ in CRITERIA],
                },
                "agreement": agreement,
                "framework_summary": framework_summary,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    print(f"Analyzed {len(sample_ids)} requirements.")
    print(f"Wrote {output_dir / 'human_eval_agreement_summary.csv'}")
    print(f"Wrote {output_dir / 'human_eval_framework_summary.csv'}")
    print(f"Wrote {output_dir / 'human_eval_item_level_scores.csv'}")
    print(f"Wrote {output_dir / 'human_eval_agreement_summary.json'}")
    return 0


def read_annotator_sheet(workbook: Any, sheet_name: str) -> dict[str, dict[str, Any]]:
    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    required = ["Sample_ID", "Case_Study", "Requirement_Set", "Requirement_Text"]
    required.extend(column for _, column in CRITERIA)
    missing = [column for column in required if column not in index]
    if missing:
        raise ValueError(f"{sheet_name} is missing columns: {missing}")

    rows: dict[str, dict[str, Any]] = {}
    for row_idx in range(2, sheet.max_row + 1):
        raw_id = sheet.cell(row_idx, index["Sample_ID"]).value
        if raw_id is None or not str(raw_id).strip():
            continue
        sample_id = str(raw_id).strip()
        if sample_id in rows:
            raise ValueError(f"{sheet_name} contains duplicate Sample_ID {sample_id}")
        row = {
            "Sample_ID": sample_id,
            "Case_Study": sheet.cell(row_idx, index["Case_Study"]).value,
            "Requirement_Set": sheet.cell(row_idx, index["Requirement_Set"]).value,
            "Requirement_Text": sheet.cell(row_idx, index["Requirement_Text"]).value,
        }
        for criterion, column in CRITERIA:
            row[criterion] = parse_score(
                sheet.cell(row_idx, index[column]).value,
                sheet_name=sheet_name,
                sample_id=sample_id,
                criterion=criterion,
            )
        rows[sample_id] = row
    return rows


def read_framework_mapping(workbook: Any) -> dict[str, str]:
    sheet_name = next((name for name in LLM_SHEET_CANDIDATES if name in workbook.sheetnames), "")
    if not sheet_name:
        return {}
    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    if "Sample_ID" not in index or "Framework" not in index:
        return {}
    mapping: dict[str, str] = {}
    for row_idx in range(2, sheet.max_row + 1):
        sample_id = sheet.cell(row_idx, index["Sample_ID"]).value
        if sample_id is None or not str(sample_id).strip():
            continue
        mapping[str(sample_id).strip()] = str(
            sheet.cell(row_idx, index["Framework"]).value or ""
        ).strip()
    return mapping


def parse_score(value: Any, *, sheet_name: str, sample_id: str, criterion: str) -> int:
    try:
        score = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Missing or non-integer score in {sheet_name}, {sample_id}, {criterion}: {value!r}"
        ) from exc
    if score not in SCORE_VALUES:
        raise ValueError(
            f"Out-of-range score in {sheet_name}, {sample_id}, {criterion}: {score}"
        )
    return score


def validate_annotators(
    annot_001: dict[str, dict[str, Any]],
    annot_002: dict[str, dict[str, Any]],
) -> list[str]:
    ids_001 = set(annot_001)
    ids_002 = set(annot_002)
    if ids_001 != ids_002:
        raise ValueError(
            "Annotator sheets contain different Sample_IDs: "
            f"only in 001={sorted(ids_001 - ids_002)}, only in 002={sorted(ids_002 - ids_001)}"
        )
    return sorted(ids_001, key=sample_sort_key)


def sample_sort_key(sample_id: str) -> tuple[int, str]:
    try:
        return (int(sample_id.split("-", 1)[1]), sample_id)
    except (IndexError, ValueError):
        return (10**9, sample_id)


def build_agreement_summary(
    annot_001: dict[str, dict[str, Any]],
    annot_002: dict[str, dict[str, Any]],
    sample_ids: list[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    all_001: list[int] = []
    all_002: list[int] = []

    for criterion, _ in CRITERIA:
        values_001 = [annot_001[sample_id][criterion] for sample_id in sample_ids]
        values_002 = [annot_002[sample_id][criterion] for sample_id in sample_ids]
        rows.append(agreement_row(criterion, values_001, values_002))
        all_001.extend(values_001)
        all_002.extend(values_002)

    rows.append(agreement_row("OVERALL_ALL_CELLS", all_001, all_002))
    return rows


def agreement_row(label: str, values_001: list[int], values_002: list[int]) -> dict[str, Any]:
    diffs = [abs(left - right) for left, right in zip(values_001, values_002, strict=True)]
    n = len(values_001)
    return {
        "criterion": label,
        "n": n,
        "mean_annot_001": round(mean(values_001), 6),
        "mean_annot_002": round(mean(values_002), 6),
        "mean_abs_diff": round(mean(diffs), 6),
        "exact_agreement_pct": round(100.0 * sum(diff == 0 for diff in diffs) / n, 6),
        "within_1_pct": round(100.0 * sum(diff <= 1 for diff in diffs) / n, 6),
        "quadratic_weighted_kappa": rounded_or_none(weighted_kappa(values_001, values_002)),
        "linear_weighted_kappa": rounded_or_none(
            weighted_kappa(values_001, values_002, weight="linear")
        ),
        "spearman": rounded_or_none(spearman(values_001, values_002)),
        "pearson": rounded_or_none(pearson(values_001, values_002)),
        "score_dist_annot_001": json.dumps(dict(sorted(Counter(values_001).items()))),
        "score_dist_annot_002": json.dumps(dict(sorted(Counter(values_002).items()))),
    }


def build_framework_summary(
    annot_001: dict[str, dict[str, Any]],
    annot_002: dict[str, dict[str, Any]],
    framework_by_id: dict[str, str],
    sample_ids: list[str],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[float]] = defaultdict(list)
    criterion_grouped: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    for sample_id in sample_ids:
        framework = framework_by_id.get(sample_id, "")
        case_study = str(annot_001[sample_id]["Case_Study"])
        for criterion, _ in CRITERIA:
            value = (annot_001[sample_id][criterion] + annot_002[sample_id][criterion]) / 2.0
            grouped[(framework, case_study)].append(value)
            grouped[(framework, "ALL")].append(value)
            criterion_grouped[(framework, case_study, criterion)].append(value)
            criterion_grouped[(framework, "ALL", criterion)].append(value)

    rows: list[dict[str, Any]] = []
    for framework, case_study in sorted(
        grouped,
        key=lambda item: (item[0], item[1] == "ALL", item[1]),
    ):
        values = grouped[(framework, case_study)]
        row: dict[str, Any] = {
            "framework": framework,
            "case_study": case_study,
            "n_requirements": int(len(values) / len(CRITERIA)),
            "n_scores": len(values),
            "mean_human_score": round(mean(values), 6),
        }
        for criterion, _ in CRITERIA:
            row[f"mean_{criterion}"] = round(
                mean(criterion_grouped[(framework, case_study, criterion)]),
                6,
            )
        rows.append(row)
    return rows


def build_item_scores(
    annot_001: dict[str, dict[str, Any]],
    annot_002: dict[str, dict[str, Any]],
    framework_by_id: dict[str, str],
    sample_ids: list[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sample_id in sample_ids:
        row: dict[str, Any] = {
            "Sample_ID": sample_id,
            "Case_Study": annot_001[sample_id]["Case_Study"],
            "Requirement_Set": annot_001[sample_id]["Requirement_Set"],
            "Framework": framework_by_id.get(sample_id, ""),
            "Requirement_Text": annot_001[sample_id]["Requirement_Text"],
        }
        human_means: list[float] = []
        for criterion, _ in CRITERIA:
            left = annot_001[sample_id][criterion]
            right = annot_002[sample_id][criterion]
            row[f"{criterion}_ANNOT_001"] = left
            row[f"{criterion}_ANNOT_002"] = right
            row[f"{criterion}_Human_Mean"] = (left + right) / 2.0
            human_means.append(row[f"{criterion}_Human_Mean"])
        row["Overall_Human_Mean"] = mean(human_means)
        rows.append(row)
    return rows


def weighted_kappa(values_001: list[int], values_002: list[int], *, weight: str = "quadratic") -> float:
    n = len(values_001)
    categories = list(SCORE_VALUES)
    index = {category: i for i, category in enumerate(categories)}
    matrix_size = len(categories)
    observed = [[0 for _ in categories] for __ in categories]
    rows = [0 for _ in categories]
    columns = [0 for _ in categories]
    for left, right in zip(values_001, values_002, strict=True):
        i = index[left]
        j = index[right]
        observed[i][j] += 1
        rows[i] += 1
        columns[j] += 1

    weights = []
    for i in range(matrix_size):
        row = []
        for j in range(matrix_size):
            distance = abs(i - j) / (matrix_size - 1)
            row.append(distance if weight == "linear" else distance * distance)
        weights.append(row)

    observed_disagreement = (
        sum(weights[i][j] * observed[i][j] for i in range(matrix_size) for j in range(matrix_size))
        / n
    )
    expected_disagreement = (
        sum(
            weights[i][j] * (rows[i] * columns[j] / n)
            for i in range(matrix_size)
            for j in range(matrix_size)
        )
        / n
    )
    if expected_disagreement == 0:
        return float("nan")
    return 1.0 - observed_disagreement / expected_disagreement


def spearman(values_001: list[int], values_002: list[int]) -> float:
    return pearson(rankdata(values_001), rankdata(values_002))


def pearson(values_001: list[float], values_002: list[float]) -> float:
    mean_001 = mean(values_001)
    mean_002 = mean(values_002)
    sum_sq_001 = sum((value - mean_001) ** 2 for value in values_001)
    sum_sq_002 = sum((value - mean_002) ** 2 for value in values_002)
    if sum_sq_001 == 0 or sum_sq_002 == 0:
        return float("nan")
    covariance = sum(
        (left - mean_001) * (right - mean_002)
        for left, right in zip(values_001, values_002, strict=True)
    )
    return covariance / math.sqrt(sum_sq_001 * sum_sq_002)


def rankdata(values: list[float]) -> list[float]:
    pairs = sorted((value, index) for index, value in enumerate(values))
    ranks = [0.0 for _ in values]
    start = 0
    while start < len(pairs):
        end = start
        while end < len(pairs) and pairs[end][0] == pairs[start][0]:
            end += 1
        average_rank = (start + 1 + end) / 2.0
        for _, index in pairs[start:end]:
            ranks[index] = average_rank
        start = end
    return ranks


def mean(values: list[float]) -> float:
    return sum(values) / len(values)


def rounded_or_none(value: float) -> float | None:
    if math.isnan(value):
        return None
    return round(value, 6)


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    raise SystemExit(main())
