#!/usr/bin/env python3
"""Build LaTeX Table C for human-vs-LLM validation."""

from __future__ import annotations

import argparse
import csv
import math
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ANNOTATOR_SHEETS = ("ANNOT-001", "ANNOT-002")
LLM_SHEET_CANDIDATES = ("LLM_Judge_Scores", "LLM-Judge(Hidden to Annotator)")
CRITERIA = (
    ("Unambiguous", "Unambiguous(1-5)", "Unambiguous"),
    ("Correctness", "Correctness(1-5)", "Correctness"),
    ("Verifiability", "Verifiability(1-5)", "Verifiability"),
    ("Set Consistency", "Set_Consistency(1-5)", "Set_Consistency"),
    ("Set Feasibility", "Set_Feasibility(1-5)", "Set_Feasibility"),
)
FRAMEWORK_ORDER = ("mare", "iredev", "quare")
SCORE_VALUES = (1, 2, 3, 4, 5)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input-xlsx",
        default="human_eval/Phase3_Human_Evaluation_Annotators.xlsx",
    )
    parser.add_argument(
        "--output-dir",
        default="human_eval/phase3_table_c_results",
    )
    args = parser.parse_args()

    workbook = load_workbook(args.input_xlsx, data_only=True)
    annotators = [read_annotator_sheet(workbook, sheet) for sheet in ANNOTATOR_SHEETS]
    sample_ids = validate_sample_ids(annotators)
    llm_scores = read_llm_scores(workbook)

    missing = [sample_id for sample_id in sample_ids if sample_id not in llm_scores]
    if missing:
        raise ValueError(f"Missing LLM scores for Sample_IDs: {missing}")

    rows = build_table_rows(annotators, llm_scores, sample_ids)
    latex = build_latex(rows)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "table_c_human_validation.csv", rows)
    (output_dir / "table_c_human_validation.tex").write_text(latex, encoding="utf-8")

    print(f"Read {len(sample_ids)} requirements from {args.input_xlsx}")
    print(f"Wrote {output_dir / 'table_c_human_validation.csv'}")
    print(f"Wrote {output_dir / 'table_c_human_validation.tex'}")
    print()
    print(latex)
    return 0


def read_annotator_sheet(workbook: Any, sheet_name: str) -> dict[str, dict[str, Any]]:
    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    required = ["Sample_ID", *(annot_col for _, annot_col, _ in CRITERIA)]
    missing = [column for column in required if column not in index]
    if missing:
        raise ValueError(f"{sheet_name} is missing columns: {missing}")

    rows: dict[str, dict[str, Any]] = {}
    for row_idx in range(2, sheet.max_row + 1):
        raw_id = sheet.cell(row_idx, index["Sample_ID"]).value
        if raw_id is None or not str(raw_id).strip():
            continue
        sample_id = str(raw_id).strip()
        row = {"Sample_ID": sample_id}
        for label, annot_col, _ in CRITERIA:
            row[label] = parse_score(sheet.cell(row_idx, index[annot_col]).value)
        rows[sample_id] = row
    return rows


def read_llm_scores(workbook: Any) -> dict[str, dict[str, Any]]:
    sheet_name = next((name for name in LLM_SHEET_CANDIDATES if name in workbook.sheetnames), "")
    if not sheet_name:
        raise ValueError(f"Could not find any LLM score sheet: {LLM_SHEET_CANDIDATES}")

    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    required = ["Sample_ID", "Framework", *(llm_col for _, _, llm_col in CRITERIA)]
    missing = [column for column in required if column not in index]
    if missing:
        raise ValueError(f"{sheet_name} is missing columns: {missing}")

    rows: dict[str, dict[str, Any]] = {}
    for row_idx in range(2, sheet.max_row + 1):
        raw_id = sheet.cell(row_idx, index["Sample_ID"]).value
        if raw_id is None or not str(raw_id).strip():
            continue
        sample_id = str(raw_id).strip()
        row = {
            "Sample_ID": sample_id,
            "Framework": str(sheet.cell(row_idx, index["Framework"]).value or "").strip().lower(),
        }
        for label, _, llm_col in CRITERIA:
            row[label] = parse_score(sheet.cell(row_idx, index[llm_col]).value)
        rows[sample_id] = row
    return rows


def validate_sample_ids(annotators: list[dict[str, dict[str, Any]]]) -> list[str]:
    base = set(annotators[0])
    for idx, annotator in enumerate(annotators[1:], start=2):
        current = set(annotator)
        if current != base:
            raise ValueError(
                f"Annotator {idx} has mismatched Sample_IDs: "
                f"only in first={sorted(base - current)}, only in annotator={sorted(current - base)}"
            )
    return sorted(base, key=sample_sort_key)


def build_table_rows(
    annotators: list[dict[str, dict[str, Any]]],
    llm_scores: dict[str, dict[str, Any]],
    sample_ids: list[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    all_human_by_framework: dict[str, list[float]] = defaultdict(list)
    all_llm_by_framework: dict[str, list[int]] = defaultdict(list)
    all_human: list[float] = []
    all_llm: list[int] = []

    for label, _, _ in CRITERIA:
        human_by_framework: dict[str, list[float]] = defaultdict(list)
        llm_by_framework: dict[str, list[int]] = defaultdict(list)
        human_values: list[float] = []
        llm_values: list[int] = []

        for sample_id in sample_ids:
            framework = llm_scores[sample_id]["Framework"]
            human_score = mean([annotator[sample_id][label] for annotator in annotators])
            llm_score = llm_scores[sample_id][label]
            human_by_framework[framework].append(human_score)
            llm_by_framework[framework].append(llm_score)
            all_human_by_framework[framework].append(human_score)
            all_llm_by_framework[framework].append(llm_score)
            human_values.append(human_score)
            llm_values.append(llm_score)
            all_human.append(human_score)
            all_llm.append(llm_score)

        rows.append(
            make_row(
                label,
                human_by_framework,
                llm_by_framework,
                human_values,
                llm_values,
            )
        )

    rows.append(
        make_row(
            "Overall",
            all_human_by_framework,
            all_llm_by_framework,
            all_human,
            all_llm,
        )
    )
    return rows


def make_row(
    criterion: str,
    human_by_framework: dict[str, list[float]],
    llm_by_framework: dict[str, list[int]],
    human_values: list[float],
    llm_values: list[int],
) -> dict[str, Any]:
    row: dict[str, Any] = {"Criterion": criterion}
    for framework in FRAMEWORK_ORDER:
        row[f"Human_{framework}"] = mean(human_by_framework[framework])
    for framework in FRAMEWORK_ORDER:
        row[f"LLM_{framework}"] = mean(llm_by_framework[framework])
    row["Spearman"] = spearman(human_values, llm_values)
    row["QWK"] = weighted_kappa([round_half_up(value) for value in human_values], llm_values)
    return row


def build_latex(rows: list[dict[str, Any]]) -> str:
    lines = [
        r"\begin{table}[t]",
        r"\centering",
        r"\caption{Human validation on the representative sample. Agreement is computed between the mean human score and the LLM judge score for the same sampled requirements; QWK uses the rounded mean human score.}",
        r"\label{tab:human-validation}",
        r"\scriptsize",
        r"\begin{tabular}{lccc ccc c}",
        r"\toprule",
        r"\multirow{2}{*}{Criterion} & \multicolumn{3}{c}{Human Eval. Mean} & \multicolumn{3}{c}{LLM Judge Mean} & \multirow{2}{*}{Agreement ($\rho$/QWK)} \\",
        r"\cmidrule(lr){2-4}\cmidrule(lr){5-7}",
        r" & MARE & iReDev & QUARE & MARE & iReDev & QUARE & \\",
        r"\midrule",
    ]
    for row in rows:
        if row["Criterion"] == "Overall":
            lines.append(r"\midrule")
        lines.append(
            " & ".join(
                [
                    latex_label(row["Criterion"]),
                    fmt(row["Human_mare"]),
                    fmt(row["Human_iredev"]),
                    fmt(row["Human_quare"]),
                    fmt(row["LLM_mare"]),
                    fmt(row["LLM_iredev"]),
                    fmt(row["LLM_quare"]),
                    f"{fmt(row['Spearman'])} / {fmt(row['QWK'])}",
                ]
            )
            + r" \\"
        )
    lines.extend(
        [
            r"\bottomrule",
            r"\end{tabular}",
            r"\end{table}",
        ]
    )
    return "\n".join(lines) + "\n"


def parse_score(value: Any) -> int:
    score = int(value)
    if score not in SCORE_VALUES:
        raise ValueError(f"Score must be 1-5, got {value!r}")
    return score


def sample_sort_key(sample_id: str) -> tuple[int, str]:
    try:
        return (int(sample_id.split("-", 1)[1]), sample_id)
    except (IndexError, ValueError):
        return (10**9, sample_id)


def mean(values: list[float]) -> float:
    return sum(values) / len(values)


def spearman(values_001: list[float], values_002: list[float]) -> float:
    return pearson(rankdata(values_001), rankdata(values_002))


def pearson(values_001: list[float], values_002: list[float]) -> float:
    mean_001 = mean(values_001)
    mean_002 = mean(values_002)
    sum_sq_001 = sum((value - mean_001) ** 2 for value in values_001)
    sum_sq_002 = sum((value - mean_002) ** 2 for value in values_002)
    if sum_sq_001 == 0 or sum_sq_002 == 0:
        return float("nan")
    covariance = sum(
        (left - mean_001) * (right - mean_002)
        for left, right in zip(values_001, values_002, strict=True)
    )
    return covariance / math.sqrt(sum_sq_001 * sum_sq_002)


def rankdata(values: list[float]) -> list[float]:
    pairs = sorted((value, index) for index, value in enumerate(values))
    ranks = [0.0 for _ in values]
    start = 0
    while start < len(pairs):
        end = start
        while end < len(pairs) and pairs[end][0] == pairs[start][0]:
            end += 1
        average_rank = (start + 1 + end) / 2.0
        for _, index in pairs[start:end]:
            ranks[index] = average_rank
        start = end
    return ranks


def weighted_kappa(values_001: list[int], values_002: list[int], *, weight: str = "quadratic") -> float:
    categories = list(SCORE_VALUES)
    index = {category: i for i, category in enumerate(categories)}
    observed = [[0 for _ in categories] for __ in categories]
    rows = [0 for _ in categories]
    columns = [0 for _ in categories]
    n = len(values_001)
    for left, right in zip(values_001, values_002, strict=True):
        observed[index[left]][index[right]] += 1
        rows[index[left]] += 1
        columns[index[right]] += 1

    weights = []
    for i in range(len(categories)):
        row = []
        for j in range(len(categories)):
            distance = abs(i - j) / (len(categories) - 1)
            row.append(distance if weight == "linear" else distance * distance)
        weights.append(row)

    observed_disagreement = (
        sum(weights[i][j] * observed[i][j] for i in range(len(categories)) for j in range(len(categories)))
        / n
    )
    expected_disagreement = (
        sum(
            weights[i][j] * (rows[i] * columns[j] / n)
            for i in range(len(categories))
            for j in range(len(categories))
        )
        / n
    )
    if expected_disagreement == 0:
        return float("nan")
    return 1.0 - observed_disagreement / expected_disagreement


def round_half_up(value: float) -> int:
    return int(math.floor(value + 0.5))


def fmt(value: float) -> str:
    if value is None or math.isnan(value):
        return "--"
    return f"{value:.2f}"


def latex_label(value: str) -> str:
    return value.replace("_", r"\_")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    raise SystemExit(main())
