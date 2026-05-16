#!/usr/bin/env python3
"""Run the LLM judge on the human-evaluation requirement sample."""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import UTC
from datetime import datetime
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from openpyxl import load_workbook
from openpyxl.comments import Comment


DEFAULT_INPUT_XLSX = "human_eval/QUARE_Human_Evaluation_Collection.xlsx"
DEFAULT_MAPPING_JSON = "human_eval/human_eval_mapping.json"
DEFAULT_OUTPUT_XLSX = "human_eval/QUARE_Human_Evaluation_Collection_with_LLM_Judge.xlsx"
DEFAULT_CACHE_JSONL = "human_eval/llm_judge_scores.jsonl"
DEFAULT_METADATA_JSON = "human_eval/llm_judge_run_metadata.json"
ANNOTATION_SHEET = "ANNOT-001"
LLM_SHEET = "LLM_Judge_Scores"
PROJECT_DEFAULT_MODEL = "gpt-4o-mini"
PROJECT_DEFAULT_TEMPERATURE = 0.7
PROJECT_DEFAULT_MAX_TOKENS = 4000
PROJECT_DEFAULT_SEED = 42
CRITERIA = (
    "Unambiguous",
    "Correctness",
    "Verifiability",
    "Set_Consistency",
    "Set_Feasibility",
)
LLM_COLUMNS = (
    "Sample_ID",
    "Case_Study",
    "Requirement_Set",
    "Framework",
    *CRITERIA,
)


class JudgeClient:
    """Small OpenAI chat client using the same config fields as OpenRE-Bench."""

    def __init__(self, *, api_key: str, model: str, base_url: str | None, timeout: float) -> None:
        from openai import OpenAI

        kwargs: dict[str, Any] = {
            "api_key": api_key,
            "timeout": timeout,
            "max_retries": 2,
        }
        if base_url:
            kwargs["base_url"] = base_url
        self._client = OpenAI(**kwargs)
        self._model = model

    def chat(
        self,
        messages: list[dict[str, str]],
        *,
        temperature: float,
        max_tokens: int,
        seed: int,
    ) -> str:
        response = self._client.chat.completions.create(
            model=self._model,
            messages=messages,
            temperature=temperature,
            max_tokens=max_tokens,
            seed=seed,
        )
        content = response.choices[0].message.content
        if not content:
            raise RuntimeError("OpenAI response did not include assistant content.")
        return content.strip()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Score human-evaluation requirements with the project LLM judge config."
    )
    parser.add_argument("--input-xlsx", default=DEFAULT_INPUT_XLSX)
    parser.add_argument("--mapping-json", default=DEFAULT_MAPPING_JSON)
    parser.add_argument("--output-xlsx", default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--cache-jsonl", default=DEFAULT_CACHE_JSONL)
    parser.add_argument("--metadata-json", default=DEFAULT_METADATA_JSON)
    parser.add_argument("--model", default=PROJECT_DEFAULT_MODEL)
    parser.add_argument("--temperature", type=float, default=PROJECT_DEFAULT_TEMPERATURE)
    parser.add_argument("--max-tokens", type=int, default=PROJECT_DEFAULT_MAX_TOKENS)
    parser.add_argument("--seed", type=int, default=PROJECT_DEFAULT_SEED)
    parser.add_argument(
        "--sync-only",
        action="store_true",
        help="Only synchronize LLM_Judge_Scores rows with the current annotation sample.",
    )
    args = parser.parse_args()

    input_xlsx = Path(args.input_xlsx)
    output_xlsx = Path(args.output_xlsx)
    cache_jsonl = Path(args.cache_jsonl)
    metadata_json = Path(args.metadata_json)

    workbook = load_workbook(input_xlsx)
    case_contexts = read_case_contexts(workbook)
    framework_by_case_set = read_framework_mapping(Path(args.mapping_json))
    requirements = read_annotation_requirements(workbook[ANNOTATION_SHEET])
    sync_llm_sheet(workbook, requirements, framework_by_case_set)

    cached_scores = read_cached_scores(cache_jsonl)
    if not args.sync_only:
        settings = load_project_openai_settings(model_override=args.model)
        client = JudgeClient(
            api_key=settings["api_key"],
            model=settings["model"],
            base_url=settings["base_url"],
            timeout=settings["timeout_seconds"],
        )
        score_groups(
            client=client,
            requirements=requirements,
            case_contexts=case_contexts,
            cached_scores=cached_scores,
            cache_jsonl=cache_jsonl,
            temperature=args.temperature,
            max_tokens=args.max_tokens,
            seed=args.seed,
        )

    apply_scores_to_workbook(workbook, requirements, cached_scores)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)

    write_metadata(
        metadata_json,
        args=args,
        input_xlsx=input_xlsx,
        output_xlsx=output_xlsx,
        cache_jsonl=cache_jsonl,
        n_requirements=len(requirements),
        n_scored=sum(1 for item in requirements if item["Sample_ID"] in cached_scores),
        model=args.model,
    )

    print(f"Requirements in current annotation sample: {len(requirements)}")
    print(f"LLM-scored requirements available: {sum(1 for item in requirements if item['Sample_ID'] in cached_scores)}")
    print(f"Wrote workbook: {output_xlsx}")
    print(f"Wrote cache: {cache_jsonl}")
    print(f"Wrote metadata: {metadata_json}")
    return 0


def read_case_contexts(workbook: Any) -> dict[str, dict[str, str]]:
    sheet = workbook["Case_Context"]
    contexts: dict[str, dict[str, str]] = {}
    for row_idx in range(4, sheet.max_row + 1):
        case_study = str(sheet.cell(row_idx, 1).value or "").strip()
        if not case_study:
            continue
        contexts[case_study] = {
            "case_description": str(sheet.cell(row_idx, 2).value or "").strip(),
            "source_requirement": str(sheet.cell(row_idx, 3).value or "").strip(),
        }
    return contexts


def read_framework_mapping(mapping_json: Path) -> dict[tuple[str, str], str]:
    if not mapping_json.exists():
        return {}
    payload = json.loads(mapping_json.read_text(encoding="utf-8"))
    mapping: dict[tuple[str, str], str] = {}
    for item in payload.get("anonymous_set_mapping", []):
        case_study = str(item.get("case_study", "")).strip()
        requirement_set = str(item.get("requirement_set", "")).strip()
        framework = str(item.get("framework", "")).strip()
        if case_study and requirement_set:
            mapping[(case_study, requirement_set)] = framework
    return mapping


def read_annotation_requirements(sheet: Any) -> list[dict[str, str]]:
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    required = ["Sample_ID", "Case_Study", "Requirement_Set", "Requirement_Text"]
    missing = [header for header in required if header not in index]
    if missing:
        raise ValueError(f"{sheet.title} is missing required columns: {missing}")

    requirements: list[dict[str, str]] = []
    seen: set[str] = set()
    for row_idx in range(2, sheet.max_row + 1):
        sample_id = str(sheet.cell(row_idx, index["Sample_ID"]).value or "").strip()
        if not sample_id:
            continue
        if sample_id in seen:
            raise ValueError(f"Duplicate Sample_ID in {sheet.title}: {sample_id}")
        seen.add(sample_id)
        requirements.append(
            {
                "Sample_ID": sample_id,
                "Case_Study": str(sheet.cell(row_idx, index["Case_Study"]).value or "").strip(),
                "Requirement_Set": str(
                    sheet.cell(row_idx, index["Requirement_Set"]).value or ""
                ).strip(),
                "Requirement_Text": str(
                    sheet.cell(row_idx, index["Requirement_Text"]).value or ""
                ).strip(),
            }
        )
    return requirements


def sync_llm_sheet(
    workbook: Any,
    requirements: list[dict[str, str]],
    framework_by_case_set: dict[tuple[str, str], str],
) -> None:
    if LLM_SHEET in workbook.sheetnames:
        sheet = workbook[LLM_SHEET]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(LLM_SHEET)

    for col_idx, header in enumerate(LLM_COLUMNS, start=1):
        sheet.cell(1, col_idx, header)

    for row_idx, requirement in enumerate(requirements, start=2):
        framework = framework_by_case_set.get(
            (requirement["Case_Study"], requirement["Requirement_Set"]),
            "",
        )
        row = {
            "Sample_ID": requirement["Sample_ID"],
            "Case_Study": requirement["Case_Study"],
            "Requirement_Set": requirement["Requirement_Set"],
            "Framework": framework,
        }
        for col_idx, header in enumerate(LLM_COLUMNS, start=1):
            sheet.cell(row_idx, col_idx, row.get(header))
    sheet.freeze_panes = "A2"


def read_cached_scores(cache_jsonl: Path) -> dict[str, dict[str, Any]]:
    scores: dict[str, dict[str, Any]] = {}
    if not cache_jsonl.exists():
        return scores
    for line in cache_jsonl.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        payload = json.loads(line)
        sample_id = str(payload.get("Sample_ID", "")).strip()
        if sample_id:
            scores[sample_id] = payload
    return scores


def load_project_openai_settings(*, model_override: str) -> dict[str, Any]:
    file_key = load_key_from_file(PROJECT_ROOT / ".api_key")
    api_key = file_key or os.environ.get("OPENAI_API_KEY") or os.environ.get("OPENAI_KEY") or ""
    api_key = api_key.strip()
    if not api_key:
        raise RuntimeError(
            "No OpenAI API key found. Create .api_key with OPENAI_API_KEY=... "
            "or set OPENAI_API_KEY in the environment."
        )
    try:
        api_key.encode("utf-8")
    except UnicodeEncodeError as exc:
        raise RuntimeError(
            "Configured OPENAI_API_KEY contains invalid non-UTF-8 characters. "
            "Create a local .api_key file with a valid key to override the current environment."
        ) from exc
    model = (model_override or os.environ.get("OPENAI_MODEL") or PROJECT_DEFAULT_MODEL).strip()
    base_url = (os.environ.get("OPENAI_BASE_URL") or "").strip() or None
    timeout = float(os.environ.get("OPENAI_TIMEOUT_SECONDS") or 180.0)
    return {
        "api_key": api_key,
        "model": model,
        "base_url": base_url,
        "timeout_seconds": max(30.0, timeout),
    }


def load_key_from_file(path: Path) -> str:
    if not path.exists():
        return ""
    for line in path.read_text(encoding="utf-8").replace("\r", "\n").split("\n"):
        value = line.strip()
        if not value or value.startswith("#"):
            continue
        if value.startswith("export "):
            value = value[7:].strip()
        if "=" not in value:
            continue
        name, raw = value.split("=", 1)
        if name.strip().upper() in {"OPENAI_API_KEY", "OPENAI_KEY"}:
            return raw.strip().strip('"').strip("'")
    return ""


def score_groups(
    *,
    client: JudgeClient,
    requirements: list[dict[str, str]],
    case_contexts: dict[str, dict[str, str]],
    cached_scores: dict[str, dict[str, Any]],
    cache_jsonl: Path,
    temperature: float,
    max_tokens: int,
    seed: int,
) -> None:
    grouped: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for requirement in requirements:
        if requirement["Sample_ID"] in cached_scores:
            continue
        grouped[(requirement["Case_Study"], requirement["Requirement_Set"])].append(requirement)

    cache_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with cache_jsonl.open("a", encoding="utf-8") as handle:
        for (case_study, requirement_set), group in sorted(grouped.items()):
            if not group:
                continue
            prompt = build_judge_prompt(
                case_study=case_study,
                requirement_set=requirement_set,
                case_context=case_contexts.get(case_study, {}),
                requirements=group,
            )
            response = client.chat(
                [
                    {
                        "role": "system",
                        "content": (
                            "You are an expert requirements-engineering evaluator. "
                            "Score requirements using ISO/IEC/IEEE 29148-style quality criteria. "
                            "Return strict JSON only."
                        ),
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=temperature,
                max_tokens=max_tokens,
                seed=seed,
            )
            parsed = parse_judge_response(response)
            expected_ids = {item["Sample_ID"] for item in group}
            returned_ids = {str(item.get("Sample_ID", "")).strip() for item in parsed}
            missing = sorted(expected_ids - returned_ids)
            extra = sorted(returned_ids - expected_ids)
            if missing or extra:
                raise ValueError(
                    f"Judge response Sample_ID mismatch for {case_study}/{requirement_set}: "
                    f"missing={missing}, extra={extra}"
                )
            for item in parsed:
                normalized = normalize_score_record(item)
                cached_scores[normalized["Sample_ID"]] = normalized
                handle.write(json.dumps(normalized, ensure_ascii=False) + "\n")
                handle.flush()
            print(f"Scored {case_study} / {requirement_set}: {len(group)} requirements")


def build_judge_prompt(
    *,
    case_study: str,
    requirement_set: str,
    case_context: dict[str, str],
    requirements: list[dict[str, str]],
) -> str:
    requirement_lines = "\n".join(
        f"- {item['Sample_ID']}: {item['Requirement_Text']}" for item in requirements
    )
    return f"""Evaluate the following requirements as a blind LLM-as-a-judge task.

Do not infer or use any framework identity. Requirement_Set is anonymous and is
provided only so Set_Consistency and Set_Feasibility can be judged against the
other requirements in the same set.

Case_Study: {case_study}
Requirement_Set: {requirement_set}
Case description: {case_context.get('case_description', '')}
Source requirement/context:
{case_context.get('source_requirement', '')}

Requirements in this same case/set:
{requirement_lines}

Score each requirement independently using integers 1-5 for:
1. Unambiguous: single clear interpretation; lower scores for vague wording,
   unclear actors, unclear conditions, or underspecified terms.
2. Correctness: reflects a real stakeholder need in the case context.
3. Verifiability: measurable/testable with observable acceptance criteria.
4. Set_Consistency: no contradiction with the other requirements in this same
   Case_Study and Requirement_Set.
5. Set_Feasibility: technically achievable and realistic for the case context
   and the same requirement set.

Return strict JSON with this exact shape and no markdown:
{{
  "scores": [
    {{
      "Sample_ID": "REQ-001",
      "Unambiguous": 1,
      "Correctness": 1,
      "Verifiability": 1,
      "Set_Consistency": 1,
      "Set_Feasibility": 1
    }}
  ]
}}
"""


def parse_judge_response(response: str) -> list[dict[str, Any]]:
    text = response.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, re.DOTALL)
        if not match:
            raise
        payload = json.loads(match.group(0))
    if isinstance(payload, dict) and isinstance(payload.get("scores"), list):
        return payload["scores"]
    if isinstance(payload, list):
        return payload
    raise ValueError("Judge response must contain a 'scores' list.")


def normalize_score_record(item: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {"Sample_ID": str(item.get("Sample_ID", "")).strip()}
    if not normalized["Sample_ID"]:
        raise ValueError(f"Judge score missing Sample_ID: {item}")
    for criterion in CRITERIA:
        value = int(item.get(criterion))
        if value not in {1, 2, 3, 4, 5}:
            raise ValueError(f"Invalid score for {normalized['Sample_ID']} / {criterion}: {value}")
        normalized[criterion] = value
    return normalized


def apply_scores_to_workbook(
    workbook: Any,
    requirements: list[dict[str, str]],
    cached_scores: dict[str, dict[str, Any]],
) -> None:
    sheet = workbook[LLM_SHEET]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: col + 1 for col, header in enumerate(headers)}
    comment = Comment(
        "LLM judge score generated by scripts/run_human_eval_llm_judge.py "
        "using the project OpenAI configuration.",
        "Codex",
    )
    for row_idx, requirement in enumerate(requirements, start=2):
        sample_id = requirement["Sample_ID"]
        score = cached_scores.get(sample_id)
        if not score:
            continue
        for criterion in CRITERIA:
            cell = sheet.cell(row_idx, index[criterion], int(score[criterion]))
            cell.comment = comment


def write_metadata(
    metadata_json: Path,
    *,
    args: argparse.Namespace,
    input_xlsx: Path,
    output_xlsx: Path,
    cache_jsonl: Path,
    n_requirements: int,
    n_scored: int,
    model: str,
) -> None:
    metadata_json.parent.mkdir(parents=True, exist_ok=True)
    metadata = {
        "created_at": datetime.now(UTC).isoformat(),
        "input_xlsx": str(input_xlsx),
        "output_xlsx": str(output_xlsx),
        "cache_jsonl": str(cache_jsonl),
        "n_requirements": n_requirements,
        "n_scored": n_scored,
        "model": model,
        "temperature": args.temperature,
        "max_tokens": args.max_tokens,
        "seed": args.seed,
        "config_note": (
            "Uses project-style OpenAI API key precedence (.api_key, OPENAI_API_KEY, "
            "OPENAI_KEY) with project default model gpt-4o-mini, temperature 0.7, "
            "max_tokens 4000 unless overridden."
        ),
    }
    metadata_json.write_text(json.dumps(metadata, indent=2, ensure_ascii=False), encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
