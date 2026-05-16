#!/usr/bin/env python3
"""Build an Excel workbook for exported Phase 3 requirements."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT_JSON = "human_eval/phase3_requirements_all.json"
DEFAULT_OUTPUT_XLSX = "human_eval/phase3_requirements_all.xlsx"
REQUIREMENT_COLUMNS = (
    "Sample_ID",
    "source_file",
    "run_id",
    "framework",
    "case_study",
    "setting",
    "seed",
    "source_requirement_id",
    "raw_requirement_text",
    "cleaned_requirement_text",
)


def main() -> int:
    parser = argparse.ArgumentParser(description="Create Phase 3 requirements Excel export.")
    parser.add_argument("--input-json", default=DEFAULT_INPUT_JSON)
    parser.add_argument("--output-xlsx", default=DEFAULT_OUTPUT_XLSX)
    args = parser.parse_args()

    records = json.loads(Path(args.input_json).read_text(encoding="utf-8"))
    if not isinstance(records, list):
        raise ValueError(f"{args.input_json} must contain a JSON list")

    workbook = Workbook()
    requirements_sheet = workbook.active
    requirements_sheet.title = "Requirements"
    populate_requirements_sheet(requirements_sheet, records)

    summary_sheet = workbook.create_sheet("Summary")
    populate_summary_sheet(summary_sheet, records)

    output_xlsx = Path(args.output_xlsx)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    print(f"Wrote {output_xlsx} with {len(records)} requirement rows")
    return 0


def populate_requirements_sheet(sheet: Any, records: list[dict[str, Any]]) -> None:
    sheet.append(list(REQUIREMENT_COLUMNS))
    for record in records:
        sheet.append([record.get(column, "") for column in REQUIREMENT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "Sample_ID": 13,
        "source_file": 72,
        "run_id": 56,
        "framework": 13,
        "case_study": 16,
        "setting": 34,
        "seed": 10,
        "source_requirement_id": 24,
        "raw_requirement_text": 92,
        "cleaned_requirement_text": 92,
    }
    for col_idx, column_name in enumerate(REQUIREMENT_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = widths[column_name]

    text_columns = {
        REQUIREMENT_COLUMNS.index("source_file") + 1,
        REQUIREMENT_COLUMNS.index("run_id") + 1,
        REQUIREMENT_COLUMNS.index("raw_requirement_text") + 1,
        REQUIREMENT_COLUMNS.index("cleaned_requirement_text") + 1,
    }
    centered_columns = {
        REQUIREMENT_COLUMNS.index("Sample_ID") + 1,
        REQUIREMENT_COLUMNS.index("framework") + 1,
        REQUIREMENT_COLUMNS.index("case_study") + 1,
        REQUIREMENT_COLUMNS.index("setting") + 1,
        REQUIREMENT_COLUMNS.index("seed") + 1,
        REQUIREMENT_COLUMNS.index("source_requirement_id") + 1,
    }
    for row_idx in range(2, sheet.max_row + 1):
        raw_text = str(
            sheet.cell(row_idx, REQUIREMENT_COLUMNS.index("raw_requirement_text") + 1).value
            or ""
        )
        cleaned_text = str(
            sheet.cell(
                row_idx,
                REQUIREMENT_COLUMNS.index("cleaned_requirement_text") + 1,
            ).value
            or ""
        )
        estimated_lines = max(2, min(12, (max(len(raw_text), len(cleaned_text)) // 90) + 2))
        sheet.row_dimensions[row_idx].height = estimated_lines * 16
        for col_idx in range(1, len(REQUIREMENT_COLUMNS) + 1):
            cell = sheet.cell(row_idx, col_idx)
            cell.alignment = Alignment(
                horizontal="center" if col_idx in centered_columns else "left",
                vertical="top",
                wrap_text=col_idx in text_columns,
            )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def populate_summary_sheet(sheet: Any, records: list[dict[str, Any]]) -> None:
    by_framework = Counter(str(record.get("framework", "")) for record in records)
    by_case = Counter(str(record.get("case_study", "")) for record in records)
    by_setting = Counter(str(record.get("setting", "")) for record in records)
    by_framework_case_setting = Counter(
        (
            str(record.get("framework", "")),
            str(record.get("case_study", "")),
            str(record.get("setting", "")),
        )
        for record in records
    )

    sheet.append(["Phase 3 Requirements Export Summary"])
    sheet.append([])
    sheet.append(["Metric", "Value"])
    sheet.append(["Total requirement rows", len(records)])
    sheet.append(["Unique run IDs", len({record.get("run_id", "") for record in records})])
    sheet.append([])

    append_counter_table(sheet, "By Framework", ["framework", "n"], by_framework)
    append_counter_table(sheet, "By Case Study", ["case_study", "n"], by_case)
    append_counter_table(sheet, "By Setting", ["setting", "n"], by_setting)

    sheet.append(["By Framework / Case Study / Setting"])
    sheet.append(["framework", "case_study", "setting", "n"])
    for (framework, case_study, setting), count in sorted(by_framework_case_setting.items()):
        sheet.append([framework, case_study, setting, count])

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = title_fill
    sheet.merge_cells("A1:D1")

    for row_idx in range(1, sheet.max_row + 1):
        row_values = [sheet.cell(row_idx, col_idx).value for col_idx in range(1, 5)]
        if row_values[0] in {
            "Metric",
            "framework",
            "case_study",
            "setting",
            "By Framework",
            "By Case Study",
            "By Setting",
            "By Framework / Case Study / Setting",
        }:
            for col_idx in range(1, 5):
                cell = sheet.cell(row_idx, col_idx)
                cell.font = Font(bold=True)
                cell.fill = header_fill

    for col_idx, width in enumerate((24, 18, 38, 12), start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.freeze_panes = "A3"


def append_counter_table(
    sheet: Any,
    title: str,
    headers: list[str],
    counter: Counter[str],
) -> None:
    sheet.append([title])
    sheet.append(headers)
    for key, count in sorted(counter.items()):
        sheet.append([key, count])
    sheet.append([])


if __name__ == "__main__":
    raise SystemExit(main())
