#!/usr/bin/env python3
"""Build a blind human-evaluation workbook for LLM-as-a-judge validation.

The script samples generated requirements from OpenRE-Bench matrix outputs,
creates an evaluator-facing Excel workbook, and writes a traceability mapping file
that links blind sample IDs back to framework/run metadata.
"""

from __future__ import annotations

import argparse
import json
import random
import re
from collections import Counter
from dataclasses import asdict
from dataclasses import dataclass
from datetime import UTC
from datetime import datetime
from pathlib import Path
from typing import Any


PHASE3_FILENAME = "phase3_integrated_kaos_model.json"
RUN_RECORD_FILENAME = "run_record.json"
DEFAULT_SETTING = "negotiation_integration_verification"
DEFAULT_SYSTEMS = ("quare", "mare", "iredev")
DEFAULT_CASE_ORDER = ("AD", "Library", "ATM", "RollCall", "Bookkeeping")
SCORE_COLUMNS = (
    "Unambiguous(1-5)",
    "Correctness(1-5)",
    "Verifiability(1-5)",
    "Set_Consistency(1-5)",
    "Set_Feasibility(1-5)",
)
CRITERIA = (
    "Unambiguous",
    "Correctness",
    "Verifiability",
    "Set Consistency",
    "Set Feasibility",
)
RUBRIC: dict[str, dict[int, str]] = {
    "Unambiguous": {
        1: "Very ambiguous; multiple plausible interpretations or unclear actors.",
        2: "Ambiguous; important terms, scope, or conditions are underspecified.",
        3: "Mostly understandable, but some vague wording or scope gaps remain.",
        4: "Clear with only minor wording ambiguity.",
        5: (
            "Single clear interpretation; avoids vague terms such as appropriate, fast, "
            "or user-friendly unless they are quantified."
        ),
    },
    "Correctness": {
        1: "Does not reflect the case or a plausible stakeholder need.",
        2: "Weakly related to the case, with major omissions or incorrect assumptions.",
        3: "Partially reflects a real stakeholder need, but needs refinement.",
        4: "Reflects a real stakeholder need with minor limitations.",
        5: "Accurately captures a real stakeholder need in the case context.",
    },
    "Verifiability": {
        1: "Not testable or measurable.",
        2: "Difficult to test; lacks clear observable outcomes.",
        3: "Somewhat testable, but acceptance criteria are incomplete or indirect.",
        4: "Testable with mostly clear acceptance criteria.",
        5: "Clearly measurable or testable with concrete acceptance criteria.",
    },
    "Set Consistency": {
        1: "Directly contradicts one or more requirements in the same case and set.",
        2: "Likely conflicts with another requirement in the same case and set.",
        3: "No obvious contradiction, but relationships or assumptions are unclear.",
        4: "Consistent with only minor tension or overlap.",
        5: "No contradiction with other requirements in the same case and set.",
    },
    "Set Feasibility": {
        1: "Technically infeasible under the case constraints.",
        2: "Likely infeasible or requires unrealistic assumptions.",
        3: "Potentially feasible but needs important clarification or constraints.",
        4: "Technically achievable with normal engineering assumptions.",
        5: "Clearly achievable and realistic for the stated system context.",
    },
}
REQUIREMENT_OR_GOAL_TYPES = {"goal", "softgoal", "requirement", "task"}
REQUIREMENT_TEXT_FIELDS = (
    "requirement_text",
    "text",
    "content",
    "description",
)
FRAMEWORK_PREFIX_LABELS = (
    "Safety",
    "Efficiency",
    "Sustainability",
    "Trustworthiness",
    "Responsibility",
    "Integrated",
    "SingleAgent",
    "SafetyAgent",
    "EfficiencyAgent",
    "GreenAgent",
    "TrustworthinessAgent",
    "ResponsibilityAgent",
    "Stakeholders",
    "Collector",
    "Modeler",
    "Checker",
    "Documenter",
    "Interviewer",
    "EndUser",
    "Deployer",
    "Analyst",
    "Archivist",
    "Reviewer",
)
FRAMEWORK_PREFIX_LABEL_RE = "|".join(re.escape(label) for label in FRAMEWORK_PREFIX_LABELS)
PIPELINE_METADATA_PATTERNS = (
    re.compile(
        r"\b(?:SafetyAgent|EfficiencyAgent|GreenAgent|TrustworthinessAgent|"
        r"ResponsibilityAgent|Stakeholders|Collector|Modeler|Checker|Documenter|"
        r"Interviewer|EndUser|Deployer|Analyst|Archivist|Reviewer)\b",
        re.IGNORECASE,
    ),
    re.compile(r"^\[?\s*conflict scenario\b", re.IGNORECASE),
    re.compile(r"^\[[^\]]+\]$", re.IGNORECASE),
    re.compile(r"^potential trade-off detected\b", re.IGNORECASE),
    re.compile(r"^external interface requirements?$", re.IGNORECASE),
    re.compile(
        r"^(?:functional|non-functional|system|software|quality) requirements?$",
        re.IGNORECASE,
    ),
    re.compile(r"^non-functional requirements and quality attributes$", re.IGNORECASE),
    re.compile(r"^functional and quality requirements$", re.IGNORECASE),
    re.compile(r"^verification and acceptance criteria$", re.IGNORECASE),
    re.compile(r"^purpose and scope$", re.IGNORECASE),
    re.compile(r"^all requirements should include\b", re.IGNORECASE),
    re.compile(r"^all requirements should remain\b", re.IGNORECASE),
    re.compile(r"^as a stakeholder\b", re.IGNORECASE),
    re.compile(r"^ensure traceability from\b", re.IGNORECASE),
    re.compile(r"^integrated single-agent objective\b", re.IGNORECASE),
    re.compile(r"^(?:srs[-\s]?\d+\s*)?(?:scope and stakeholders)$", re.IGNORECASE),
    re.compile(r"^(?:srs[-\s]?\d+\s*)?(?:purpose and scope)$", re.IGNORECASE),
    re.compile(
        r"^(?:srs[-\s]?\d+\s*)?(?:functional and quality requirements)$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^(?:srs[-\s]?\d+\s*)?(?:non-functional requirements"
        r"(?: and quality attributes)?)$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^(?:srs[-\s]?\d+\s*)?(?:verification and acceptance criteria)$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^(?:srs[-\s]?\d+\s*)?(?:external interface requirements?)$",
        re.IGNORECASE,
    ),
    re.compile(r"^srs[-\s]?\d*\s*$", re.IGNORECASE),
    re.compile(r"^(?:the system shall\s+)?(?:entity|relation)\s*:", re.IGNORECASE),
    re.compile(r"^(?:draft requirement|interview record|check report item)\b", re.IGNORECASE),
    re.compile(r"^(?:closure|finding|response|deploy constraint)\s*:", re.IGNORECASE),
    re.compile(r"^(?:selected model|requirement model)\s*:", re.IGNORECASE),
    re.compile(r"\bfinding:\s*potential trade-off\b", re.IGNORECASE),
    re.compile(r"^resolved$", re.IGNORECASE),
)
FRAMEWORK_PREFIX_PATTERNS = (
    re.compile(
        rf"^(?:{FRAMEWORK_PREFIX_LABEL_RE})\s+"
        r"(?:Output|Requirement|Goal)(?:\s+\d+)?\s*:\s*",
        re.IGNORECASE,
    ),
    re.compile(
        rf"^(?:{FRAMEWORK_PREFIX_LABEL_RE})\s+objective\s+for\s+"
        r"[A-Za-z0-9_-]+\s*:\s*",
        re.IGNORECASE,
    ),
    re.compile(r"^Negotiated\s+Integrated\s+Goal\s*:\s*", re.IGNORECASE),
    re.compile(
        r"^Integrated\s+single-agent\s+objective\s+for\s+"
        r"[A-Za-z0-9_-]+\s*:\s*",
        re.IGNORECASE,
    ),
    re.compile(
        r"^The\s+system\s+shall\s+ensure\s+"
        r"(?:safety|efficiency|sustainability|trustworthiness|responsibility)"
        r"\s*\([^)]*\)\s*:\s*,?\s*",
        re.IGNORECASE,
    ),
    re.compile(
        r"^The\s+system\s+shall\s+(?="
        r"(?:Draft requirement|Interview record|Check report item|Response|"
        r"Deploy constraint|Closure|Finding|Selected model|Requirement model|"
        r"User requirement|Operating environment|System requirement|Review finding|"
        r"SR[-\s]?\d+|UR[-\s]?\d+|REQ[-\s]?\d+|SRS[-\s]?\d+|"
        r"Entity\s*:|Relation\s*:|As a stakeholder))",
        re.IGNORECASE,
    ),
    re.compile(r"^The\s+system\s+shall\s+[-•]\s*", re.IGNORECASE),
    re.compile(r"^(?:Draft\s+requirement|Interview\s+record)\s+\d+\s*:\s*", re.IGNORECASE),
    re.compile(r"^Check\s+report\s+item\s+\d+\s*:\s*", re.IGNORECASE),
    re.compile(
        r"^(?:Review\s+finding|User\s+requirement|Operating\s+environment|"
        r"System\s+requirement|Requirement\s+model)\s+\d+\s*:\s*",
        re.IGNORECASE,
    ),
    re.compile(r"^(?:SR|UR|REQ)\s*[- ]?\d+\s*:\s*", re.IGNORECASE),
    re.compile(r"^SRS\s*[- ]?\d+\s*:\s*", re.IGNORECASE),
    re.compile(r"^The\s+system\s+shall\s+satisfy\s*:\s*", re.IGNORECASE),
    re.compile(
        r"^The\s+system\s+shall\s+(?="
        r"(?:It|A|An|Each|For each|Recorded transactions|Neither|Curvature constraints|"
        r"Non-convex obstacle constraints|Optimization|Students|Members|Librarians|"
        r"Teachers|Users|Clients|The application|A checking-account|A bank client)\b)",
        re.IGNORECASE,
    ),
    re.compile(r"^(?:Response|Deploy\s+constraint|Closure|Finding)\s*:\s*", re.IGNORECASE),
    re.compile(r"^(?:Selected\s+model|Requirement\s+model)\s*:\s*", re.IGNORECASE),
    re.compile(r"^As a stakeholder of [^,]{1,80},\s*I need\s+", re.IGNORECASE),
    re.compile(r"^[-•]\s*", re.IGNORECASE),
)


@dataclass(frozen=True)
class RequirementCandidate:
    framework: str
    case_study: str
    setting: str
    seed: int
    run_id: str
    source_file: str
    source_requirement_id: str
    source_requirement_name: str
    source_element_type: str
    quality_attribute: str
    stakeholder: str
    raw_requirement_text: str
    requirement_text: str


@dataclass(frozen=True)
class SampleRecord:
    sample_id: str
    case_study: str
    requirement_set: str
    requirement_text: str
    framework: str
    setting: str
    seed: int
    run_id: str
    source_file: str
    source_requirement_id: str
    source_requirement_name: str
    source_element_type: str
    quality_attribute: str
    stakeholder: str
    raw_requirement_text: str
    display_order: float


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    matrix_output_dir = Path(args.matrix_output_dir)
    runs_dir = Path(args.runs_dir) if args.runs_dir else matrix_output_dir / "runs"
    cases_dir = Path(args.cases_dir)
    output_xlsx = Path(args.output_xlsx)
    mapping_json = Path(args.mapping_json)

    systems = parse_csv(args.systems) or list(DEFAULT_SYSTEMS)
    case_filter = parse_csv(args.cases)
    case_order = parse_csv(args.case_order) or list(DEFAULT_CASE_ORDER)

    case_contexts = load_case_contexts(cases_dir)
    candidates = collect_candidates(
        runs_dir=runs_dir,
        systems=systems,
        cases=case_filter,
        setting=args.setting,
        include_goals=bool(args.include_goals) or not bool(args.exclude_goals),
    )
    if not candidates:
        raise SystemExit(
            "No requirement candidates found. Run the comparison matrix first, "
            "or pass --matrix-output-dir/--runs-dir pointing to generated runs."
        )

    samples, set_mapping, shortfalls = sample_requirements(
        candidates=candidates,
        samples_per_pair=args.samples_per_pair,
        seed=args.seed,
        case_order=case_order,
    )
    if not samples:
        raise SystemExit("Sampling produced zero records.")

    write_workbook(
        output_xlsx=output_xlsx,
        samples=samples,
        case_contexts=case_contexts,
        include_set_column=not args.no_set_column,
        case_order=case_order,
    )
    write_mapping(
        mapping_json=mapping_json,
        samples=samples,
        set_mapping=set_mapping,
        shortfalls=shortfalls,
        args=args,
        matrix_output_dir=matrix_output_dir,
        runs_dir=runs_dir,
        output_xlsx=output_xlsx,
    )

    print(f"Human evaluation workbook written: {output_xlsx}")
    print(f"Private sample mapping written: {mapping_json}")
    print(f"Sampled requirements: {len(samples)}")
    if shortfalls:
        print("Sampling shortfalls:")
        for item in shortfalls:
            print(
                "- "
                f"{item['framework']} / {item['case_study']}: "
                f"requested {item['requested']}, available {item['available']}"
            )
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Create a blind Excel workbook for human ISO 29148 evaluation."
    )
    parser.add_argument(
        "--matrix-output-dir",
        default="experiment_outputs/mare-iredev-quare",
        help="Matrix output directory containing runs/.",
    )
    parser.add_argument(
        "--runs-dir",
        default="",
        help="Optional direct path to the runs directory; overrides --matrix-output-dir/runs.",
    )
    parser.add_argument(
        "--cases-dir",
        default="data/case_studies",
        help="Directory containing *_input.json case files for evaluator context.",
    )
    parser.add_argument(
        "--output-xlsx",
        default="human_eval/human_eval_annotation.xlsx",
        help="Path for the evaluator-facing workbook.",
    )
    parser.add_argument(
        "--mapping-json",
        default="human_eval/human_eval_mapping.json",
        help="Path for the private Sample_ID to framework mapping JSON.",
    )
    parser.add_argument(
        "--samples-per-pair",
        type=int,
        default=10,
        help="Requirements sampled per framework x case_study pair.",
    )
    parser.add_argument("--seed", type=int, default=42, help="Sampling seed.")
    parser.add_argument(
        "--setting",
        default=DEFAULT_SETTING,
        help=(
            "Run setting to sample. Use the full-pipeline setting by default; "
            "pass 'all' to sample every setting."
        ),
    )
    parser.add_argument(
        "--systems",
        default=",".join(DEFAULT_SYSTEMS),
        help="Comma-separated frameworks/systems to sample.",
    )
    parser.add_argument(
        "--cases",
        default="",
        help="Optional comma-separated case studies to include.",
    )
    parser.add_argument(
        "--case-order",
        default=",".join(DEFAULT_CASE_ORDER),
        help="Comma-separated case-study display order in the workbook.",
    )
    parser.add_argument(
        "--include-goals",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--exclude-goals",
        action="store_true",
        help="Exclude top-level goals; by default true requirements and goals are eligible.",
    )
    parser.add_argument(
        "--no-set-column",
        action="store_true",
        help="Omit the blind Requirement_Set column from Annotation.",
    )
    return parser


def parse_csv(raw: str) -> list[str]:
    return [item.strip() for item in raw.split(",") if item.strip()]


def load_case_contexts(cases_dir: Path) -> dict[str, dict[str, str]]:
    contexts: dict[str, dict[str, str]] = {}
    if not cases_dir.exists():
        return contexts

    for path in sorted(cases_dir.glob("*_input.json")):
        payload = load_json(path)
        if not isinstance(payload, dict):
            continue
        case_name = str(payload.get("case_name", "")).strip()
        if not case_name:
            continue
        contexts[case_name] = {
            "case_description": str(payload.get("case_description", "")).strip(),
            "requirement": str(payload.get("requirement", "")).strip(),
        }
    return contexts


def collect_candidates(
    *,
    runs_dir: Path,
    systems: list[str],
    cases: list[str],
    setting: str,
    include_goals: bool,
) -> list[RequirementCandidate]:
    system_filter = {item.lower() for item in systems}
    case_filter = set(cases)
    setting_filter = setting.strip()
    candidates: list[RequirementCandidate] = []

    if not runs_dir.exists():
        return candidates

    for run_record_path in sorted(runs_dir.glob(f"*/{RUN_RECORD_FILENAME}")):
        run_record = load_json(run_record_path)
        if not isinstance(run_record, dict):
            continue

        framework = str(run_record.get("system", "")).strip().lower()
        case_study = str(run_record.get("case_id", "")).strip()
        run_setting = str(run_record.get("setting", "")).strip()
        if system_filter and framework not in system_filter:
            continue
        if case_filter and case_study not in case_filter:
            continue
        if setting_filter and setting_filter.lower() != "all" and run_setting != setting_filter:
            continue

        phase3_path = resolve_phase3_path(run_record, run_record_path.parent)
        if phase3_path is None or not phase3_path.exists():
            continue

        phase3 = load_json(phase3_path)
        elements = phase3.get("gsn_elements", []) if isinstance(phase3, dict) else []
        if not isinstance(elements, list):
            continue

        for element in elements:
            if not isinstance(element, dict):
                continue
            raw_requirement_text = format_raw_requirement_text(element)
            requirement_text = format_requirement_text(element)
            if not is_requirement_like(
                element,
                requirement_text=requirement_text,
                include_goals=include_goals,
            ):
                continue

            normalized = normalize_text(requirement_text)
            if not normalized:
                continue

            properties = element.get("properties", {})
            if not isinstance(properties, dict):
                properties = {}

            candidates.append(
                RequirementCandidate(
                    framework=framework,
                    case_study=case_study,
                    setting=run_setting,
                    seed=to_int(run_record.get("seed"), default=0),
                    run_id=str(run_record.get("run_id", "")).strip(),
                    source_file=str(phase3_path),
                    source_requirement_id=str(element.get("id", "")).strip(),
                    source_requirement_name=str(element.get("name", "")).strip(),
                    source_element_type=(
                        str(properties.get("original_kaos_type", "")).strip()
                        or str(element.get("gsn_type", "")).strip()
                    ),
                    quality_attribute=str(element.get("quality_attribute", "")).strip(),
                    stakeholder=str(element.get("stakeholder", "")).strip(),
                    raw_requirement_text=raw_requirement_text,
                    requirement_text=requirement_text,
                )
            )

    return candidates


def resolve_phase3_path(run_record: dict[str, Any], run_dir: Path) -> Path | None:
    artifact_paths = run_record.get("artifact_paths", {})
    if isinstance(artifact_paths, dict):
        raw_path = str(artifact_paths.get(PHASE3_FILENAME, "")).strip()
        if raw_path:
            candidate = Path(raw_path)
            if candidate.exists():
                return candidate
            relative_candidate = run_dir / candidate.name
            if relative_candidate.exists():
                return relative_candidate

    artifacts_dir = str(run_record.get("artifacts_dir", "")).strip()
    if artifacts_dir:
        candidate = Path(artifacts_dir) / PHASE3_FILENAME
        if candidate.exists():
            return candidate

    candidate = run_dir / PHASE3_FILENAME
    if candidate.exists():
        return candidate
    return None


def is_requirement_like(
    element: dict[str, Any],
    *,
    requirement_text: str,
    include_goals: bool,
) -> bool:
    if not is_substantive_requirement_text(requirement_text):
        return False
    properties = element.get("properties", {})
    if not isinstance(properties, dict):
        properties = {}
    original_type = str(properties.get("original_kaos_type", "")).strip().lower()
    hierarchy_level = to_int(element.get("hierarchy_level"), default=1)
    gsn_type = str(element.get("gsn_type", "")).strip().lower()

    if original_type in {"goal", "softgoal"}:
        return include_goals
    if original_type in REQUIREMENT_OR_GOAL_TYPES:
        return True
    if gsn_type == "goal":
        return include_goals
    if gsn_type == "strategy" and hierarchy_level >= 2:
        return True
    return False


def format_raw_requirement_text(element: dict[str, Any]) -> str:
    """Use semantic text fields only; do not prepend GSN names or agent labels."""

    for field_name in REQUIREMENT_TEXT_FIELDS:
        value = str(element.get(field_name, "")).strip()
        if value:
            return value
    return ""


def format_requirement_text(element: dict[str, Any]) -> str:
    cleaned = strip_framework_identifying_prefixes(format_raw_requirement_text(element))
    return first_requirement_segment(cleaned)


def strip_framework_identifying_prefixes(text: str) -> str:
    cleaned = normalize_spaces(text)
    for _ in range(16):
        previous = cleaned
        cleaned = cleaned.strip()
        for pattern in FRAMEWORK_PREFIX_PATTERNS:
            cleaned = pattern.sub("", cleaned).strip()
        cleaned = re.sub(
            r"^(?:The\s+system\s+shall\s+){2,}",
            "The system shall ",
            cleaned,
            flags=re.IGNORECASE,
        )
        cleaned = re.sub(
            r"^The\s+system\s+shall\s+(?=The\s+(?:system|ATM)\s+(?:shall|must|should|will)\b)",
            "",
            cleaned,
            flags=re.IGNORECASE,
        )
        cleaned = re.sub(r"^[,;:\)\]\s]+", "", cleaned)
        cleaned = normalize_spaces(cleaned)
        if cleaned == previous:
            break
    return cleaned.strip()


def first_requirement_segment(text: str) -> str:
    """Return the first complete requirement-like segment without truncating it."""

    cleaned = normalize_spaces(text)
    if not cleaned:
        return cleaned

    bullet_match = re.search(r"\s[-•]\s+(?=[A-Z0-9])", cleaned)
    sentence_index = first_sentence_boundary(cleaned)
    repeated_requirement_index = first_repeated_requirement_boundary(cleaned)

    boundary_indexes = (
        bullet_match.start() if bullet_match else -1,
        sentence_index,
        repeated_requirement_index,
    )
    candidates = [index for index in boundary_indexes if index > 0]
    if not candidates:
        return cleaned
    return cleaned[: min(candidates)].strip(" ,;:-")


def first_sentence_boundary(text: str) -> int:
    protected, placeholders = protect_sentence_abbreviations(text)
    match = re.search(r"(?<=[.!?])\s+(?=[A-Z0-9])", protected)
    if not match:
        return -1
    restored_prefix = restore_sentence_abbreviations(protected[: match.start()], placeholders)
    return len(restored_prefix)


def first_repeated_requirement_boundary(text: str) -> int:
    pattern = re.compile(
        r"\s+(?=(?:The\s+(?:system|application|user|client|vehicle)|"
        r"(?:It|Users?|Members?|Clients?|Students?|Teachers?|Librarians?|"
        r"Administrators?|Recorded transactions)\s+"
        r"(?:shall|must|should|will|can|are|required)|"
        r"Neither|For each|If|Upon|"
        r"A\s+(?:bank client|checking-account|bookkeeping system|"
        r"library management system|roll call system))\b)",
        re.IGNORECASE,
    )
    for match in pattern.finditer(text):
        if match.start() >= 32:
            return match.start()
    return -1


def protect_sentence_abbreviations(text: str) -> tuple[str, dict[str, str]]:
    placeholders: dict[str, str] = {}
    protected = text
    abbreviations = (
        "e.g.",
        "E.g.",
        "i.e.",
        "I.e.",
        "etc.",
        "vs.",
        "No.",
        "Fig.",
        "Sec.",
    )
    for index, abbreviation in enumerate(abbreviations):
        placeholder = f"__ABBR_{index}__"
        placeholders[placeholder] = abbreviation
        protected = protected.replace(abbreviation, placeholder)
    return protected, placeholders


def restore_sentence_abbreviations(text: str, placeholders: dict[str, str]) -> str:
    restored = text
    for placeholder, abbreviation in placeholders.items():
        restored = restored.replace(placeholder, abbreviation)
    return restored


def is_substantive_requirement_text(text: str) -> bool:
    normalized = normalize_spaces(text).strip()
    if len(normalized) < 24:
        return False
    if len(normalized.split()) < 5:
        return False
    if normalized.endswith(":"):
        return False
    if re.search(r"(?:,\s*)?(?:and|or)$", normalized, re.IGNORECASE):
        return False
    if re.search(r"\b(?:a|an|the|to|of|with|for)$", normalized, re.IGNORECASE):
        return False
    if "..." in normalized or "…" in normalized:
        return False
    if re.search(r"\(\s*e\s*,", normalized, re.IGNORECASE):
        return False
    if "[" in normalized or "]" in normalized:
        return False
    if re.match(r"^[^\w]", normalized):
        return False
    if re.match(r"^[a-z]", normalized) and not re.match(r"^(?:a|an)\s+", normalized, re.IGNORECASE):
        return False
    if has_unbalanced_parentheses(normalized):
        return False
    if not has_requirement_or_goal_cue(normalized):
        return False
    if re.search(r"\b(?:SR|UR|REQ|SRS)\s*[- ]?\d+\s*:", normalized, re.IGNORECASE):
        return False
    if re.search(r"\bThe\s+system\s+shall\s+(?:Entity|SRS|Draft requirement)\b", normalized, re.IGNORECASE):
        return False
    for pattern in PIPELINE_METADATA_PATTERNS:
        if pattern.search(normalized):
            return False
    return True


def has_requirement_or_goal_cue(text: str) -> bool:
    if re.search(
        r"\b(?:shall|must|should|will|can|may|able to|required to|is required to|"
        r"are required to|needs? to|operates?)\b",
        text,
        re.IGNORECASE,
    ):
        return True
    if re.match(r"^The\s+system\s+operates\b", text, re.IGNORECASE):
        return True
    return False


def has_unbalanced_parentheses(text: str) -> bool:
    depth = 0
    for char in text:
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
            if depth < 0:
                return True
    return depth != 0


def sample_requirements(
    *,
    candidates: list[RequirementCandidate],
    samples_per_pair: int,
    seed: int,
    case_order: list[str],
) -> tuple[list[SampleRecord], list[dict[str, str]], list[dict[str, Any]]]:
    rng = random.Random(seed)
    grouped: dict[tuple[str, str], list[RequirementCandidate]] = {}
    for candidate in candidates:
        grouped.setdefault((candidate.framework, candidate.case_study), []).append(candidate)

    selected: list[RequirementCandidate] = []
    shortfalls: list[dict[str, Any]] = []
    for key in sorted(grouped, key=lambda item: (case_sort_key(item[1], case_order), item[0])):
        group = list(grouped[key])
        picked = choose_group_sample(group, samples_per_pair, rng)
        take = len(picked)
        selected.extend(picked)
        if take < samples_per_pair:
            framework, case_study = key
            shortfalls.append(
                {
                    "framework": framework,
                    "case_study": case_study,
                    "requested": samples_per_pair,
                    "available": count_unique_requirements(group),
                    "raw_candidates": len(group),
                    "reason": "Fewer unique cleaned requirements were available.",
                }
            )

    set_by_case_framework, set_mapping = assign_blind_sets(selected, rng, case_order)

    sample_id_order = list(selected)
    rng.shuffle(sample_id_order)
    sample_ids = {
        candidate_key(candidate): f"REQ-{index:03d}"
        for index, candidate in enumerate(sample_id_order, start=1)
    }

    records: list[SampleRecord] = []
    for candidate in selected:
        records.append(
            SampleRecord(
                sample_id=sample_ids[candidate_key(candidate)],
                case_study=candidate.case_study,
                requirement_set=set_by_case_framework[
                    (candidate.case_study, candidate.framework)
                ],
                requirement_text=candidate.requirement_text,
                framework=candidate.framework,
                setting=candidate.setting,
                seed=candidate.seed,
                run_id=candidate.run_id,
                source_file=candidate.source_file,
                source_requirement_id=candidate.source_requirement_id,
                source_requirement_name=candidate.source_requirement_name,
                source_element_type=candidate.source_element_type,
                quality_attribute=candidate.quality_attribute,
                stakeholder=candidate.stakeholder,
                raw_requirement_text=candidate.raw_requirement_text,
                display_order=rng.random(),
            )
        )

    records.sort(
        key=lambda item: (
            case_sort_key(item.case_study, case_order),
            item.requirement_set,
            item.display_order,
        )
    )
    return records, set_mapping, shortfalls


def choose_group_sample(
    group: list[RequirementCandidate],
    samples_per_pair: int,
    rng: random.Random,
) -> list[RequirementCandidate]:
    """Sample up to the requested count after exact deduplication by cleaned text."""

    shuffled = list(group)
    rng.shuffle(shuffled)

    by_text: dict[str, list[RequirementCandidate]] = {}
    for candidate in shuffled:
        by_text.setdefault(normalize_text(candidate.requirement_text), []).append(candidate)

    unique_first = [rng.choice(candidates) for candidates in by_text.values()]
    rng.shuffle(unique_first)
    return unique_first[:samples_per_pair]


def count_unique_requirements(group: list[RequirementCandidate]) -> int:
    return len({normalize_text(candidate.requirement_text) for candidate in group})


def assign_blind_sets(
    selected: list[RequirementCandidate],
    rng: random.Random,
    case_order: list[str],
) -> tuple[dict[tuple[str, str], str], list[dict[str, str]]]:
    frameworks_by_case: dict[str, set[str]] = {}
    for candidate in selected:
        frameworks_by_case.setdefault(candidate.case_study, set()).add(candidate.framework)

    set_by_case_framework: dict[tuple[str, str], str] = {}
    mapping: list[dict[str, str]] = []
    for case_study in sorted(frameworks_by_case, key=lambda item: case_sort_key(item, case_order)):
        frameworks = sorted(frameworks_by_case[case_study])
        rng.shuffle(frameworks)
        for index, framework in enumerate(frameworks, start=1):
            requirement_set = set_label(index)
            set_by_case_framework[(case_study, framework)] = requirement_set
            mapping.append(
                {
                    "case_study": case_study,
                    "requirement_set": requirement_set,
                    "framework": framework,
                }
            )
    return set_by_case_framework, mapping


def write_workbook(
    *,
    output_xlsx: Path,
    samples: list[SampleRecord],
    case_contexts: dict[str, dict[str, str]],
    include_set_column: bool,
    case_order: list[str],
) -> None:
    try:
        import pandas as pd
        from openpyxl.styles import Alignment
        from openpyxl.styles import Border
        from openpyxl.styles import Font
        from openpyxl.styles import PatternFill
        from openpyxl.styles import Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise SystemExit(
            "This script requires pandas and openpyxl. Install them with "
            "`uv sync --all-groups` or `python -m pip install pandas openpyxl`."
        ) from exc

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    rows = [annotation_row(item, include_set_column=include_set_column) for item in samples]
    annotation_columns = annotation_column_order(include_set_column=include_set_column)

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=annotation_columns).to_excel(
            writer,
            sheet_name="Annotation",
            index=False,
        )
        workbook = writer.book
        rubric_sheet = workbook.create_sheet("Rubric", 0)
        populate_rubric_sheet(rubric_sheet)
        style_rubric_sheet(rubric_sheet)
        case_context_sheet = workbook.create_sheet("Case Context", 1)
        populate_case_context_sheet(case_context_sheet, case_contexts, case_order)
        style_case_context_sheet(case_context_sheet)

        annotation_sheet = workbook["Annotation"]
        score_columns = list(SCORE_COLUMNS)
        style_annotation_sheet(
            annotation_sheet,
            annotation_columns=annotation_columns,
            score_columns=score_columns,
            include_set_column=include_set_column,
        )

        # Keep imports used inside the writer context visible to static checkers.
        _ = Alignment, Border, Font, PatternFill, Side, get_column_letter, DataValidation


def populate_rubric_sheet(
    sheet: Any,
) -> None:
    sheet.title = "Rubric"
    rows: list[list[Any]] = [
        ["Human Evaluation Rubric for ISO/IEC/IEEE 29148 Requirement Quality"],
        [],
        ["Instructions"],
        [
            "Score each requirement independently on every criterion using integers 1-5. "
            "Use the full scale when justified."
        ],
        [
            "The workbook is blind: Requirement_Set is an anonymous set identifier, not a "
            "framework name."
        ],
        [
            "For Set Consistency, compare only against other requirements with the same "
            "Case_Study and Requirement_Set."
        ],
        [
            "For Correctness and Set Feasibility, use the Case Context sheet. Add Notes "
            "when a score depends on an assumption."
        ],
        [],
        ["Criterion", "Score", "Anchor Description"],
    ]
    for criterion in CRITERIA:
        anchors = RUBRIC[criterion]
        for score in range(1, 6):
            rows.append([criterion if score == 1 else "", score, anchors[score]])

    for row in rows:
        sheet.append(row)


def populate_case_context_sheet(
    sheet: Any,
    case_contexts: dict[str, dict[str, str]],
    case_order: list[str],
) -> None:
    sheet.title = "Case Context"
    sheet.append(["Case Context for Human Evaluation"])
    sheet.append([])
    sheet.append(["Case_Study", "Case_Description", "Source Requirement"])
    ordered_cases = sorted(
        case_contexts,
        key=lambda item: case_sort_key(item, case_order),
    )
    for case_name in ordered_cases:
        context = case_contexts[case_name]
        sheet.append(
            [
                case_name,
                context.get("case_description", ""),
                context.get("requirement", ""),
            ]
        )


def style_rubric_sheet(sheet: Any) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill

    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="E2F0D9")

    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 88
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = title_fill
    sheet.merge_cells("A1:C1")

    for row_idx in range(1, sheet.max_row + 1):
        first_value = sheet.cell(row_idx, 1).value
        if first_value == "Instructions":
            for col_idx in range(1, 4):
                cell = sheet.cell(row_idx, col_idx)
                cell.font = Font(bold=True)
                cell.fill = section_fill
        if first_value in {"Criterion", "Case_Study"}:
            for col_idx in range(1, 4):
                cell = sheet.cell(row_idx, col_idx)
                cell.font = Font(bold=True)
                cell.fill = header_fill

    for row_idx in range(4, 8):
        sheet.row_dimensions[row_idx].height = 36


def style_case_context_sheet(sheet: Any) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="E2F0D9")

    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 104
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = title_fill
    sheet.merge_cells("A1:C1")
    for col_idx in range(1, 4):
        cell = sheet.cell(3, col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_idx in range(4, sheet.max_row + 1):
        sheet.row_dimensions[row_idx].height = 96


def style_annotation_sheet(
    sheet: Any,
    *,
    annotation_columns: list[str],
    score_columns: list[str],
    include_set_column: bool,
) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.styles import Border
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    header_fill = PatternFill("solid", fgColor="1F4E78")
    case_fills = (
        PatternFill("solid", fgColor="FFFFFF"),
        PatternFill("solid", fgColor="F7FBFF"),
    )
    set_fill = PatternFill("solid", fgColor="F2F2F2")
    top_border = Border(top=Side(style="thin", color="7F7F7F"))

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    widths = {
        "Sample_ID": 13,
        "Case_Study": 16,
        "Requirement_Set": 18,
        "Requirement_Text": 92,
        "Notes": 32,
    }
    for score_col in score_columns:
        widths[score_col] = 18
    for index, column_name in enumerate(annotation_columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(column_name, 16)

    score_indexes = [
        annotation_columns.index(score_col) + 1
        for score_col in score_columns
        if score_col in annotation_columns
    ]
    requirement_text_col = annotation_columns.index("Requirement_Text") + 1
    case_col = annotation_columns.index("Case_Study") + 1
    set_col = (
        annotation_columns.index("Requirement_Set") + 1
        if include_set_column and "Requirement_Set" in annotation_columns
        else None
    )

    validation = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    validation.error = "Enter an integer score from 1 to 5."
    validation.errorTitle = "Invalid score"
    validation.prompt = "Choose a score from 1 to 5."
    validation.promptTitle = "ISO 29148 score"
    sheet.add_data_validation(validation)
    for score_index in score_indexes:
        col_letter = get_column_letter(score_index)
        validation.add(f"{col_letter}2:{col_letter}{sheet.max_row}")

    previous_case = None
    fill_index = 0
    for row_idx in range(2, sheet.max_row + 1):
        case_value = sheet.cell(row_idx, case_col).value
        if case_value != previous_case:
            fill_index += 1
            previous_case = case_value
            for col_idx in range(1, len(annotation_columns) + 1):
                sheet.cell(row_idx, col_idx).border = top_border
        fill = case_fills[fill_index % len(case_fills)]
        for col_idx in range(1, len(annotation_columns) + 1):
            cell = sheet.cell(row_idx, col_idx)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if set_col is not None:
            sheet.cell(row_idx, set_col).fill = set_fill
        sheet.cell(row_idx, requirement_text_col).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        text_length = len(str(sheet.cell(row_idx, requirement_text_col).value or ""))
        estimated_lines = max(3, (text_length // 90) + 2)
        sheet.row_dimensions[row_idx].height = min(409, estimated_lines * 18)

    for score_index in score_indexes:
        for row_idx in range(2, sheet.max_row + 1):
            sheet.cell(row_idx, score_index).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )


def annotation_column_order(*, include_set_column: bool) -> list[str]:
    columns = ["Sample_ID", "Case_Study"]
    if include_set_column:
        columns.append("Requirement_Set")
    columns.extend(
        [
            "Requirement_Text",
            *SCORE_COLUMNS,
            "Notes",
        ]
    )
    return columns


def annotation_row(item: SampleRecord, *, include_set_column: bool) -> dict[str, Any]:
    row: dict[str, Any] = {
        "Sample_ID": item.sample_id,
        "Case_Study": item.case_study,
        "Requirement_Text": item.requirement_text,
        "Unambiguous(1-5)": "",
        "Correctness(1-5)": "",
        "Verifiability(1-5)": "",
        "Set_Consistency(1-5)": "",
        "Set_Feasibility(1-5)": "",
        "Notes": "",
    }
    if include_set_column:
        row["Requirement_Set"] = item.requirement_set
    return row


def write_mapping(
    *,
    mapping_json: Path,
    samples: list[SampleRecord],
    set_mapping: list[dict[str, str]],
    shortfalls: list[dict[str, Any]],
    args: argparse.Namespace,
    matrix_output_dir: Path,
    runs_dir: Path,
    output_xlsx: Path,
) -> None:
    mapping_json.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "metadata": {
            "created_at": datetime.now(UTC).isoformat(),
            "seed": args.seed,
            "samples_per_pair": args.samples_per_pair,
            "setting": args.setting,
            "systems": parse_csv(args.systems) or list(DEFAULT_SYSTEMS),
            "cases": parse_csv(args.cases),
            "matrix_output_dir": str(matrix_output_dir),
            "runs_dir": str(runs_dir),
            "annotation_workbook": str(output_xlsx),
            "total_samples": len(samples),
            "score_columns": list(SCORE_COLUMNS),
            "deduplication_policy": (
                "Exact duplicates are removed within each framework x case-study "
                "candidate pool after text cleaning. The workbook samples up to "
                "the requested count of unique cleaned requirements per pair and "
                "does not fill shortfalls with duplicate text."
            ),
            "sample_source_element_types": dict(
                sorted(Counter(sample.source_element_type for sample in samples).items())
            ),
            "blind_set_note": (
                "Requirement_Set is anonymous within each case study. "
                "Use this mapping only after collecting human annotations."
            ),
        },
        "anonymous_set_mapping": set_mapping,
        "sampling_shortfalls": shortfalls,
        "samples": [
            {
                key: value
                for key, value in asdict(sample).items()
                if key not in {"display_order"}
            }
            for sample in sorted(samples, key=lambda item: item.sample_id)
        ],
    }
    mapping_json.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def candidate_key(candidate: RequirementCandidate) -> tuple[str, str, str, str, str, str]:
    return (
        candidate.framework,
        candidate.case_study,
        candidate.run_id,
        candidate.source_requirement_id,
        candidate.source_requirement_name,
        candidate.requirement_text,
    )


def case_sort_key(case_study: str, case_order: list[str]) -> tuple[int, str]:
    order_map = {name: index for index, name in enumerate(case_order)}
    return (order_map.get(case_study, len(order_map)), case_study)


def set_label(index: int) -> str:
    if index <= 0:
        raise ValueError("Set index must be positive.")
    value = index
    letters: list[str] = []
    while value:
        value, remainder = divmod(value - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return f"Set_{''.join(reversed(letters))}"


def normalize_text(text: str) -> str:
    return normalize_spaces(text).lower()


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def to_int(value: Any, *, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def load_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


if __name__ == "__main__":
    raise SystemExit(main())
