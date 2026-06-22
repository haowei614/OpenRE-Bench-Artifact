#!/usr/bin/env python3
"""Build a blind human-evaluation workbook sampled from Phase 3 outputs."""

from __future__ import annotations

import argparse
import json
import random
import sys
from collections import Counter
from collections import defaultdict
from datetime import UTC
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from build_human_eval_workbook import RUBRIC
from build_human_eval_workbook import normalize_text


DEFAULT_INPUT_JSON = "human_eval/phase3_requirements_all.json"
DEFAULT_CASES_DIR = "data/case_studies"
DEFAULT_OUTPUT_XLSX = "human_eval/Phase3_Human_Evaluation_Collection.xlsx"
DEFAULT_MAPPING_JSON = "human_eval/phase3_human_eval_mapping.json"
DEFAULT_SUMMARY_JSON = "human_eval/phase3_human_eval_sample_summary.json"
DEFAULT_SETTING = "negotiation_integration_verification"
DEFAULT_SAMPLES_PER_PAIR = 10
DEFAULT_SEED = 42
CASE_ORDER = ("AD", "Library", "ATM", "RollCall", "Bookkeeping")
FRAMEWORK_ORDER = ("quare", "mare", "iredev")
ANNOTATOR_SHEETS = ("ANNOT-001", "ANNOT-002")
CRITERIA = (
    "Unambiguous",
    "Correctness",
    "Verifiability",
    "Set_Consistency",
    "Set_Feasibility",
)
ANNOTATION_COLUMNS = (
    "Sample_ID",
    "Case_Study",
    "Requirement_Set",
    "Requirement_Text",
    "Unambiguous(1-5)",
    "Correctness(1-5)",
    "Verifiability(1-5)",
    "Set_Consistency(1-5)",
    "Set_Feasibility(1-5)",
    "Notes",
)
LLM_COLUMNS = (
    "Sample_ID",
    "Case_Study",
    "Requirement_Set",
    "Framework",
    *CRITERIA,
)
SCORE_COLUMNS = ANNOTATION_COLUMNS[4:9]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Create a blind human-evaluation collection from Phase 3 outputs."
    )
    parser.add_argument("--input-json", default=DEFAULT_INPUT_JSON)
    parser.add_argument("--cases-dir", default=DEFAULT_CASES_DIR)
    parser.add_argument("--output-xlsx", default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--mapping-json", default=DEFAULT_MAPPING_JSON)
    parser.add_argument("--summary-json", default=DEFAULT_SUMMARY_JSON)
    parser.add_argument("--setting", default=DEFAULT_SETTING)
    parser.add_argument("--samples-per-pair", type=int, default=DEFAULT_SAMPLES_PER_PAIR)
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    args = parser.parse_args()

    phase3_records = load_phase3_records(Path(args.input_json), setting=args.setting)
    samples, shortfalls = sample_records(
        phase3_records,
        samples_per_pair=args.samples_per_pair,
        seed=args.seed,
    )
    workbook = build_workbook(samples, load_case_contexts(Path(args.cases_dir)))

    output_xlsx = Path(args.output_xlsx)
    mapping_json = Path(args.mapping_json)
    summary_json = Path(args.summary_json)
    for path in (output_xlsx, mapping_json, summary_json):
        path.parent.mkdir(parents=True, exist_ok=True)

    workbook.save(output_xlsx)
    mapping_payload = build_mapping_payload(samples, shortfalls, args)
    mapping_json.write_text(
        json.dumps(mapping_payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    summary_payload = build_summary_payload(samples, shortfalls, args)
    summary_json.write_text(
        json.dumps(summary_payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    print(f"Wrote {output_xlsx}")
    print(f"Wrote {mapping_json}")
    print(f"Wrote {summary_json}")
    print(f"Sampled occurrence rows: {len(samples)}")
    print(f"Unique visible requirement texts: {count_unique_texts(samples)}")
    return 0


def load_phase3_records(input_json: Path, *, setting: str) -> list[dict[str, Any]]:
    payload = json.loads(input_json.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError(f"{input_json} must contain a JSON list.")
    records = [
        record
        for record in payload
        if str(record.get("setting", "")).strip() == setting
        and str(record.get("framework", "")).strip() in FRAMEWORK_ORDER
        and str(record.get("case_study", "")).strip() in CASE_ORDER
        and str(record.get("cleaned_requirement_text", "")).strip()
    ]
    if not records:
        raise ValueError(f"No Phase 3 records found for setting={setting!r}.")
    return records


def sample_records(
    records: list[dict[str, Any]],
    *,
    samples_per_pair: int,
    seed: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rng = random.Random(seed)
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        grouped[(str(record["framework"]), str(record["case_study"]))].append(record)

    selected: list[dict[str, Any]] = []
    shortfalls: list[dict[str, Any]] = []
    for framework in FRAMEWORK_ORDER:
        for case_study in CASE_ORDER:
            group = grouped.get((framework, case_study), [])
            by_text: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for record in group:
                by_text[normalize_text(str(record["cleaned_requirement_text"]))].append(record)
            representatives = [rng.choice(items) for items in by_text.values()]
            rng.shuffle(representatives)
            picked = representatives[:samples_per_pair]
            selected.extend(picked)
            if len(representatives) < samples_per_pair:
                shortfalls.append(
                    {
                        "framework": framework,
                        "case_study": case_study,
                        "requested": samples_per_pair,
                        "available_unique": len(representatives),
                        "raw_records": len(group),
                        "reason": "Fewer unique cleaned Phase 3 requirement texts were available.",
                    }
                )

    set_by_case_framework = assign_blind_sets(selected, rng)
    duplicate_group_by_text = assign_duplicate_groups(selected)

    id_order = list(selected)
    rng.shuffle(id_order)
    sample_id_by_key = {
        source_key(record): f"P3-{index:03d}"
        for index, record in enumerate(id_order, start=1)
    }

    samples: list[dict[str, Any]] = []
    for record in selected:
        sample = {
            **record,
            "phase3_export_sample_id": record.get("Sample_ID", ""),
            "Sample_ID": sample_id_by_key[source_key(record)],
            "Requirement_Set": set_by_case_framework[
                (str(record["case_study"]), str(record["framework"]))
            ],
            "duplicate_text_group": duplicate_group_by_text[
                normalize_text(str(record["cleaned_requirement_text"]))
            ],
            "display_order": rng.random(),
        }
        samples.append(sample)

    samples.sort(
        key=lambda item: (
            case_sort_key(str(item["case_study"])),
            str(item["Requirement_Set"]),
            float(item["display_order"]),
        )
    )
    return samples, shortfalls


def assign_blind_sets(
    selected: list[dict[str, Any]],
    rng: random.Random,
) -> dict[tuple[str, str], str]:
    frameworks_by_case: dict[str, set[str]] = defaultdict(set)
    for record in selected:
        frameworks_by_case[str(record["case_study"])].add(str(record["framework"]))

    mapping: dict[tuple[str, str], str] = {}
    for case_study in sorted(frameworks_by_case, key=case_sort_key):
        frameworks = sorted(frameworks_by_case[case_study])
        rng.shuffle(frameworks)
        for index, framework in enumerate(frameworks):
            mapping[(case_study, framework)] = f"Set_{chr(ord('A') + index)}"
    return mapping


def assign_duplicate_groups(selected: list[dict[str, Any]]) -> dict[str, str]:
    counts = Counter(normalize_text(str(item["cleaned_requirement_text"])) for item in selected)
    groups: dict[str, str] = {}
    duplicate_index = 1
    for text in sorted(counts):
        if counts[text] > 1:
            groups[text] = f"DUP-{duplicate_index:03d}"
            duplicate_index += 1
        else:
            groups[text] = ""
    return groups


def build_workbook(samples: list[dict[str, Any]], case_contexts: dict[str, dict[str, str]]) -> Workbook:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    populate_instructions_sheet(instructions)
    style_instructions_sheet(instructions)

    context_sheet = workbook.create_sheet("Case_Context")
    populate_case_context_sheet(context_sheet, case_contexts)
    style_case_context_sheet(context_sheet)

    for sheet_name in ANNOTATOR_SHEETS:
        sheet = workbook.create_sheet(sheet_name)
        populate_annotator_sheet(sheet, samples)
        style_annotator_sheet(sheet)

    llm_sheet = workbook.create_sheet("LLM_Judge_Scores")
    populate_llm_sheet(llm_sheet, samples)
    style_llm_sheet(llm_sheet)
    llm_sheet.sheet_state = "hidden"
    workbook.active = 0
    return workbook


def populate_instructions_sheet(sheet: Any) -> None:
    sheet.append(["Human Evaluation Rubric for ISO/IEC/IEEE 29148 Requirement Quality"])
    sheet.append([])
    sheet.append(["Instructions"])
    sheet.append(
        [
            "Score each requirement independently on every criterion using integers 1-5. "
            "Use the full scale when justified."
        ]
    )
    sheet.append(
        [
            "The workbook is blind: Requirement_Set is an anonymous set identifier, not a "
            "framework name."
        ]
    )
    sheet.append(
        [
            "For Set Consistency, compare only against other requirements with the same "
            "Case_Study and Requirement_Set."
        ]
    )
    sheet.append(
        [
            "For Correctness and Set Feasibility, use the Case_Context sheet. Add Notes "
            "when a score depends on an assumption."
        ]
    )
    sheet.append([])
    sheet.append(["Criterion", "Score", "Anchor Description"])
    for criterion in ("Unambiguous", "Correctness", "Verifiability", "Set Consistency", "Set Feasibility"):
        for score in range(1, 6):
            sheet.append([criterion if score == 1 else "", score, RUBRIC[criterion][score]])


def populate_case_context_sheet(sheet: Any, case_contexts: dict[str, dict[str, str]]) -> None:
    sheet.append(["Case Context for Human Evaluation"])
    sheet.append([])
    sheet.append(["Case_Study", "Case_Description", "Source Requirement"])
    for case_study in CASE_ORDER:
        context = case_contexts.get(case_study, {})
        sheet.append(
            [
                case_study,
                context.get("case_description", ""),
                context.get("requirement", ""),
            ]
        )


def populate_annotator_sheet(sheet: Any, samples: list[dict[str, Any]]) -> None:
    sheet.append(list(ANNOTATION_COLUMNS))
    for sample in samples:
        sheet.append(
            [
                sample["Sample_ID"],
                sample["case_study"],
                sample["Requirement_Set"],
                sample["cleaned_requirement_text"],
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )


def populate_llm_sheet(sheet: Any, samples: list[dict[str, Any]]) -> None:
    sheet.append(list(LLM_COLUMNS))
    fill_comment = Comment(
        "TO BE FILLED: run LLM judge on these Phase 3 requirements using the same "
        "evaluation prompt from the experiment.",
        "Codex",
    )
    for sample in samples:
        sheet.append(
            [
                sample["Sample_ID"],
                sample["case_study"],
                sample["Requirement_Set"],
                sample["framework"],
                None,
                None,
                None,
                None,
                None,
            ]
        )
    for col_idx, header in enumerate(LLM_COLUMNS, start=1):
        if header in CRITERIA:
            sheet.cell(1, col_idx).comment = fill_comment


def style_instructions_sheet(sheet: Any) -> None:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 92
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = title_fill
    sheet.merge_cells("A1:C1")
    for row_idx in range(1, sheet.max_row + 1):
        first = sheet.cell(row_idx, 1).value
        if first == "Instructions":
            for col_idx in range(1, 4):
                sheet.cell(row_idx, col_idx).font = Font(bold=True)
                sheet.cell(row_idx, col_idx).fill = section_fill
        if first == "Criterion":
            for col_idx in range(1, 4):
                sheet.cell(row_idx, col_idx).font = Font(bold=True)
                sheet.cell(row_idx, col_idx).fill = header_fill


def style_case_context_sheet(sheet: Any) -> None:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 36
    sheet.column_dimensions["C"].width = 104
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = title_fill
    sheet.merge_cells("A1:C1")
    for col_idx in range(1, 4):
        sheet.cell(3, col_idx).font = Font(bold=True)
        sheet.cell(3, col_idx).fill = header_fill
    for row_idx in range(4, sheet.max_row + 1):
        sheet.row_dimensions[row_idx].height = 96


def style_annotator_sheet(sheet: Any) -> None:
    style_tabular_sheet(sheet, ANNOTATION_COLUMNS, requirement_column="Requirement_Text")
    add_score_validation(sheet, SCORE_COLUMNS)


def style_llm_sheet(sheet: Any) -> None:
    style_tabular_sheet(sheet, LLM_COLUMNS, requirement_column=None)
    add_score_validation(sheet, CRITERIA)


def style_tabular_sheet(
    sheet: Any,
    columns: tuple[str, ...],
    *,
    requirement_column: str | None,
) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    case_fills = (
        PatternFill("solid", fgColor="FFFFFF"),
        PatternFill("solid", fgColor="F7FBFF"),
    )
    top_border = Border(top=Side(style="thin", color="7F7F7F"))
    header_index = {header: index + 1 for index, header in enumerate(columns)}
    score_columns = set(SCORE_COLUMNS) | set(CRITERIA)
    requirement_col_idx = header_index.get(requirement_column) if requirement_column else None
    case_col_idx = header_index.get("Case_Study")

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "Sample_ID": 12,
        "Case_Study": 16,
        "Requirement_Set": 16,
        "Requirement_Text": 74,
        "Framework": 14,
        "Notes": 34,
    }
    for criterion in score_columns:
        widths[criterion] = 18
    for col_idx, header in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 16)

    previous_case = None
    fill_index = 0
    for row_idx in range(2, sheet.max_row + 1):
        case_value = sheet.cell(row_idx, case_col_idx).value if case_col_idx else None
        if case_value != previous_case:
            fill_index += 1
            previous_case = case_value
            for col_idx in range(1, len(columns) + 1):
                sheet.cell(row_idx, col_idx).border = top_border
        fill = case_fills[fill_index % len(case_fills)]
        for col_idx, header in enumerate(columns, start=1):
            cell = sheet.cell(row_idx, col_idx)
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="center" if header in score_columns else "left",
                vertical="center" if header in score_columns else "top",
                wrap_text=col_idx == requirement_col_idx,
            )
        if requirement_col_idx:
            text = str(sheet.cell(row_idx, requirement_col_idx).value or "")
            sheet.row_dimensions[row_idx].height = min(240, max(42, (len(text) // 74 + 2) * 18))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def add_score_validation(sheet: Any, score_columns: tuple[str, ...]) -> None:
    headers = [sheet.cell(1, col_idx).value for col_idx in range(1, sheet.max_column + 1)]
    index = {header: col_idx + 1 for col_idx, header in enumerate(headers)}
    validation = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    validation.error = "Enter an integer score from 1 to 5."
    validation.errorTitle = "Invalid score"
    validation.prompt = "Choose a score from 1 to 5."
    validation.promptTitle = "ISO 29148 score"
    sheet.add_data_validation(validation)
    for column_name in score_columns:
        if column_name not in index:
            continue
        col_letter = get_column_letter(index[column_name])
        validation.add(f"{col_letter}2:{col_letter}{sheet.max_row}")


def load_case_contexts(cases_dir: Path) -> dict[str, dict[str, str]]:
    contexts: dict[str, dict[str, str]] = {}
    for path in sorted(cases_dir.glob("*_input.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        contexts[str(payload.get("case_name", "")).strip()] = {
            "case_description": str(payload.get("case_description", "")).strip(),
            "requirement": str(payload.get("requirement", "")).strip(),
        }
    return contexts


def build_mapping_payload(
    samples: list[dict[str, Any]],
    shortfalls: list[dict[str, Any]],
    args: argparse.Namespace,
) -> dict[str, Any]:
    set_mapping = sorted(
        {
            (
                str(sample["case_study"]),
                str(sample["Requirement_Set"]),
                str(sample["framework"]),
            )
            for sample in samples
        }
    )
    return {
        "metadata": {
            "created_at": datetime.now(UTC).isoformat(),
            "source": args.input_json,
            "setting": args.setting,
            "seed": args.seed,
            "samples_per_pair": args.samples_per_pair,
            "total_samples": len(samples),
            "unique_visible_requirement_texts": count_unique_texts(samples),
            "duplicate_occurrence_rows": len(samples) - count_unique_texts(samples),
            "note": (
                "Sample_ID rows are blind occurrence samples from Phase 3 final outputs. "
                "Exact duplicate cleaned texts may appear across frameworks; duplicate_text_group "
                "identifies those occurrences for optional downstream deduplication."
            ),
        },
        "anonymous_set_mapping": [
            {
                "case_study": case_study,
                "requirement_set": requirement_set,
                "framework": framework,
            }
            for case_study, requirement_set, framework in set_mapping
        ],
        "sampling_shortfalls": shortfalls,
        "samples": [mapping_record(sample) for sample in sorted(samples, key=lambda item: item["Sample_ID"])],
    }


def build_summary_payload(
    samples: list[dict[str, Any]],
    shortfalls: list[dict[str, Any]],
    args: argparse.Namespace,
) -> dict[str, Any]:
    return {
        "metadata": {
            "created_at": datetime.now(UTC).isoformat(),
            "source": args.input_json,
            "setting": args.setting,
            "seed": args.seed,
            "samples_per_pair": args.samples_per_pair,
        },
        "total_samples": len(samples),
        "unique_visible_requirement_texts": count_unique_texts(samples),
        "duplicate_occurrence_rows": len(samples) - count_unique_texts(samples),
        "by_framework": dict(sorted(Counter(str(sample["framework"]) for sample in samples).items())),
        "by_case_study": dict(sorted(Counter(str(sample["case_study"]) for sample in samples).items())),
        "by_framework_case": [
            {"framework": framework, "case_study": case_study, "n": count}
            for (framework, case_study), count in sorted(
                Counter((str(sample["framework"]), str(sample["case_study"])) for sample in samples).items()
            )
        ],
        "duplicate_groups": duplicate_group_summary(samples),
        "sampling_shortfalls": shortfalls,
    }


def mapping_record(sample: dict[str, Any]) -> dict[str, Any]:
    return {
        "Sample_ID": sample["Sample_ID"],
        "Requirement_Set": sample["Requirement_Set"],
        "framework": sample["framework"],
        "case_study": sample["case_study"],
        "setting": sample["setting"],
        "seed": sample["seed"],
        "run_id": sample["run_id"],
        "source_file": sample["source_file"],
        "phase3_export_sample_id": sample.get("phase3_export_sample_id", ""),
        "source_requirement_id": sample["source_requirement_id"],
        "raw_requirement_text": sample["raw_requirement_text"],
        "cleaned_requirement_text": sample["cleaned_requirement_text"],
        "duplicate_text_group": sample["duplicate_text_group"],
    }


def duplicate_group_summary(samples: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for sample in samples:
        group_id = str(sample.get("duplicate_text_group", ""))
        if group_id:
            grouped[group_id].append(sample)
    return [
        {
            "duplicate_text_group": group_id,
            "n": len(items),
            "sample_ids": [str(item["Sample_ID"]) for item in sorted(items, key=lambda row: str(row["Sample_ID"]))],
            "frameworks": sorted({str(item["framework"]) for item in items}),
            "case_studies": sorted({str(item["case_study"]) for item in items}),
            "text": str(items[0]["cleaned_requirement_text"]),
        }
        for group_id, items in sorted(grouped.items())
    ]


def count_unique_texts(samples: list[dict[str, Any]]) -> int:
    return len({normalize_text(str(sample["cleaned_requirement_text"])) for sample in samples})


def source_key(record: dict[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        str(record["source_file"]),
        str(record["run_id"]),
        str(record["source_requirement_id"]),
        str(record["framework"]),
        str(record["cleaned_requirement_text"]),
    )


def case_sort_key(case_study: str) -> tuple[int, str]:
    try:
        return (CASE_ORDER.index(case_study), case_study)
    except ValueError:
        return (len(CASE_ORDER), case_study)


if __name__ == "__main__":
    raise SystemExit(main())
